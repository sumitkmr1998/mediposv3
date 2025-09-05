from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Union
import uuid
from datetime import datetime, timedelta
from enum import Enum
import subprocess
import json
import shutil
import tarfile
import tempfile
import hashlib
from passlib.context import CryptContext
from jose import JWTError, jwt
from functools import wraps


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Authentication configuration
SECRET_KEY = os.environ.get("SECRET_KEY", "your-secret-key-here-change-this-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app without a prefix
app = FastAPI(title="MediPOS API", description="Comprehensive Pharmacy Management System")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Enums
class PaymentMethod(str, Enum):
    CASH = "cash"
    CARD = "card"
    UPI = "upi"
    CREDIT = "credit"

class StockStatus(str, Enum):
    IN_STOCK = "in_stock"
    LOW_STOCK = "low_stock"
    OUT_OF_STOCK = "out_of_stock"

class TransactionType(str, Enum):
    SALE = "sale"
    PURCHASE = "purchase"
    RETURN = "return"
    ADJUSTMENT = "adjustment"
    REFUND = "refund"


# Define Models
class Settings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    general: dict = {}
    opd_paper: dict = {}
    printer: dict = {}
    telegram: dict = {}
    alerts: dict = {}
    custom_templates: dict = {}  # New field for custom templates
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class SettingsUpdate(BaseModel):
    general: Optional[dict] = None
    opd_paper: Optional[dict] = None
    printer: Optional[dict] = None
    telegram: Optional[dict] = None
    alerts: Optional[dict] = None
    custom_templates: Optional[dict] = None  # New field for custom templates
class Medicine(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    generic_name: Optional[str] = None
    manufacturer: Optional[str] = None
    batch_number: Optional[str] = None
    expiry_date: Optional[datetime] = None
    purchase_price: float
    selling_price: float
    stock_quantity: int = 0
    minimum_stock_level: int = 10
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class MedicineCreate(BaseModel):
    name: str
    generic_name: Optional[str] = None
    manufacturer: Optional[str] = None
    batch_number: Optional[str] = None
    expiry_date: Optional[datetime] = None
    purchase_price: float
    selling_price: float
    stock_quantity: int = 0
    minimum_stock_level: int = 10
    description: Optional[str] = None

class MedicineUpdate(BaseModel):
    name: Optional[str] = None
    generic_name: Optional[str] = None
    manufacturer: Optional[str] = None
    batch_number: Optional[str] = None
    expiry_date: Optional[datetime] = None
    purchase_price: Optional[float] = None
    selling_price: Optional[float] = None
    stock_quantity: Optional[int] = None
    minimum_stock_level: Optional[int] = None
    description: Optional[str] = None

class Patient(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    date_of_birth: Optional[datetime] = None
    gender: Optional[str] = None
    emergency_contact: Optional[str] = None
    medical_history: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class PatientCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    date_of_birth: Optional[datetime] = None
    gender: Optional[str] = None
    emergency_contact: Optional[str] = None
    medical_history: Optional[str] = None

class PatientUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    date_of_birth: Optional[datetime] = None
    gender: Optional[str] = None
    emergency_contact: Optional[str] = None
    medical_history: Optional[str] = None

class SaleItem(BaseModel):
    medicine_id: str
    medicine_name: str
    quantity: int
    unit_price: float
    total_price: float

class Sale(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax_amount: float = 0.0
    discount_amount: float = 0.0
    total_amount: float
    payment_method: PaymentMethod
    created_at: datetime = Field(default_factory=datetime.utcnow)

class SaleCreate(BaseModel):
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax_amount: float = 0.0
    discount_amount: float = 0.0
    total_amount: float
    payment_method: PaymentMethod

class StockMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    medicine_id: str
    medicine_name: str
    transaction_type: TransactionType
    quantity: int
    unit_price: float
    total_value: float
    reference_id: Optional[str] = None  # Sale ID, Purchase ID, etc.
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

# Permission definitions
class Permission(BaseModel):
    # Medicine permissions
    medicines_view: bool = True
    medicines_add: bool = False
    medicines_edit: bool = False
    medicines_delete: bool = False
    
    # Patient permissions
    patients_view: bool = True
    patients_add: bool = False
    patients_edit: bool = False
    patients_delete: bool = False
    
    # Sales permissions
    sales_view: bool = True
    sales_add: bool = False
    sales_edit: bool = False
    sales_delete: bool = False
    sales_refund: bool = False
    
    # Doctor permissions
    doctors_view: bool = True
    doctors_add: bool = False
    doctors_edit: bool = False
    doctors_delete: bool = False
    
    # OPD permissions
    opd_view: bool = True
    opd_add: bool = False
    opd_edit: bool = False
    opd_delete: bool = False
    
    # Analytics permissions
    analytics_view: bool = False
    analytics_export: bool = False
    
    # Settings permissions
    settings_view: bool = False
    settings_edit: bool = False
    
    # User management permissions
    users_view: bool = False
    users_add: bool = False
    users_edit: bool = False
    users_delete: bool = False
    
    # Backup permissions
    backup_create: bool = False
    backup_restore: bool = False
    backup_delete: bool = False

# Default permission sets for roles
def get_default_permissions(role: str) -> Permission:
    if role == "admin":
        # Admin has all permissions
        return Permission(
            medicines_view=True, medicines_add=True, medicines_edit=True, medicines_delete=True,
            patients_view=True, patients_add=True, patients_edit=True, patients_delete=True,
            sales_view=True, sales_add=True, sales_edit=True, sales_delete=True, sales_refund=True,
            doctors_view=True, doctors_add=True, doctors_edit=True, doctors_delete=True,
            opd_view=True, opd_add=True, opd_edit=True, opd_delete=True,
            analytics_view=True, analytics_export=True,
            settings_view=True, settings_edit=True,
            users_view=True, users_add=True, users_edit=True, users_delete=True,
            backup_create=True, backup_restore=True, backup_delete=True
        )
    elif role == "manager":
        # Manager has most permissions except user management and critical settings
        return Permission(
            medicines_view=True, medicines_add=True, medicines_edit=True, medicines_delete=False,
            patients_view=True, patients_add=True, patients_edit=True, patients_delete=False,
            sales_view=True, sales_add=True, sales_edit=True, sales_delete=False, sales_refund=True,
            doctors_view=True, doctors_add=True, doctors_edit=True, doctors_delete=False,
            opd_view=True, opd_add=True, opd_edit=True, opd_delete=False,
            analytics_view=True, analytics_export=True,
            settings_view=True, settings_edit=False,
            users_view=False, users_add=False, users_edit=False, users_delete=False,
            backup_create=True, backup_restore=False, backup_delete=False
        )
    else:  # staff
        # Staff has basic permissions
        return Permission(
            medicines_view=True, medicines_add=False, medicines_edit=False, medicines_delete=False,
            patients_view=True, patients_add=True, patients_edit=True, patients_delete=False,
            sales_view=True, sales_add=True, sales_edit=False, sales_delete=False, sales_refund=False,
            doctors_view=True, doctors_add=False, doctors_edit=False, doctors_delete=False,
            opd_view=True, opd_add=True, opd_edit=False, opd_delete=False,
            analytics_view=False, analytics_export=False,
            settings_view=False, settings_edit=False,
            users_view=False, users_add=False, users_edit=False, users_delete=False,
            backup_create=False, backup_restore=False, backup_delete=False
        )

class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    full_name: str
    role: str = "staff"  # admin, manager, staff
    permissions: Permission = Field(default_factory=lambda: get_default_permissions("staff"))
    is_active: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class UserInDB(User):
    hashed_password: str

class UserCreate(BaseModel):
    username: str
    email: str
    full_name: str
    password: str
    role: str = "staff"
    permissions: Optional[Permission] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    full_name: Optional[str] = None
    role: Optional[str] = None
    permissions: Optional[Permission] = None
    is_active: Optional[bool] = None

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class TokenData(BaseModel):
    username: Optional[str] = None

class Doctor(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    specialization: str
    qualification: str
    license_number: str
    phone: Optional[str] = None
    email: Optional[str] = None
    clinic_name: Optional[str] = None
    clinic_address: Optional[str] = None
    consultation_fee: Optional[float] = None
    signature_url: Optional[str] = None  # URL to doctor's signature image
    is_active: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class DoctorCreate(BaseModel):
    name: str
    specialization: str
    qualification: str
    license_number: str
    phone: Optional[str] = None
    email: Optional[str] = None
    clinic_name: Optional[str] = None
    clinic_address: Optional[str] = None
    consultation_fee: Optional[float] = None
    signature_url: Optional[str] = None

class DoctorUpdate(BaseModel):
    name: Optional[str] = None
    specialization: Optional[str] = None
    qualification: Optional[str] = None
    license_number: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    clinic_name: Optional[str] = None
    clinic_address: Optional[str] = None
    consultation_fee: Optional[float] = None
    signature_url: Optional[str] = None
    is_active: Optional[bool] = None

class OPDPrescription(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    doctor_id: str
    patient_id: str
    date: datetime = Field(default_factory=datetime.utcnow)
    consultation_fee: Optional[float] = None
    prescription_notes: Optional[str] = None  # Handwritten prescription area
    next_visit_date: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class OPDPrescriptionCreate(BaseModel):
    doctor_id: str
    patient_id: str
    consultation_fee: Optional[float] = None
    prescription_notes: Optional[str] = None
    next_visit_date: Optional[datetime] = None

class Return(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    original_sale_id: str
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax_amount: float = 0.0
    discount_amount: float = 0.0
    total_amount: float
    reason: Optional[str] = None
    refund_method: PaymentMethod
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ReturnCreate(BaseModel):
    original_sale_id: str
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax_amount: float = 0.0
    discount_amount: float = 0.0
    total_amount: float
    reason: Optional[str] = None
    refund_method: PaymentMethod

class CustomTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    html: str
    css: str
    category: str = "custom"  # custom, system, shared
    is_public: bool = False
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class CustomTemplateCreate(BaseModel):
    name: str
    description: Optional[str] = None
    html: str
    css: str
    category: str = "custom"
    is_public: bool = False

class CustomTemplateUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    html: Optional[str] = None
    css: Optional[str] = None
    category: Optional[str] = None
    is_public: Optional[bool] = None

class ReturnCreate(BaseModel):
    original_sale_id: str
    patient_id: Optional[str] = None
    patient_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax_amount: float = 0.0
    discount_amount: float = 0.0
    total_amount: float
    reason: Optional[str] = None
    refund_method: PaymentMethod

# Authentication Functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_user(username: str):
    user = await db.users.find_one({"username": username})
    if user:
        # Handle permissions field conversion from dict to Permission object
        if "permissions" in user and isinstance(user["permissions"], dict):
            user["permissions"] = Permission(**user["permissions"])
        elif "permissions" not in user:
            # For backward compatibility, set default permissions
            user["permissions"] = get_default_permissions(user.get("role", "staff"))
        return UserInDB(**user)
    return None

async def authenticate_user(username: str, password: str):
    user = await get_user(username)
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    return user

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    
    user = await get_user(username=token_data.username)
    if user is None:
        raise credentials_exception
    return user

async def get_current_active_user(current_user: UserInDB = Depends(get_current_user)):
    if not current_user.is_active:
        raise HTTPException(status_code=400, detail="Inactive user")
    return current_user

def require_role(allowed_roles: List[str]):
    def decorator(func):
        @wraps(func)
        async def wrapper(*args, current_user: UserInDB = Depends(get_current_active_user), **kwargs):
            if current_user.role not in allowed_roles:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Not enough permissions"
                )
            return await func(*args, current_user=current_user, **kwargs)
        return wrapper
    return decorator

def require_permission(permission_name: str):
    """Decorator to check if user has specific permission"""
    def decorator(func):
        @wraps(func)
        async def wrapper(*args, current_user: UserInDB = Depends(get_current_active_user), **kwargs):
            # Admin always has all permissions
            if current_user.role == "admin":
                return await func(*args, current_user=current_user, **kwargs)
            
            # Check if user has the specific permission
            user_permissions = current_user.permissions
            if not hasattr(user_permissions, permission_name):
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"Invalid permission: {permission_name}"
                )
            
            if not getattr(user_permissions, permission_name):
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"Missing permission: {permission_name}"
                )
            
            return await func(*args, current_user=current_user, **kwargs)
        return wrapper
    return decorator

def check_permission(user: UserInDB, permission_name: str) -> bool:
    """Helper function to check if user has specific permission"""
    if user.role == "admin":
        return True
    
    user_permissions = user.permissions
    if not hasattr(user_permissions, permission_name):
        return False
    
    return getattr(user_permissions, permission_name)

# Basic routes
@api_router.get("/")
async def root():
    return {"message": "MediPOS API - Pharmacy Management System"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow()}


# Authentication APIs
@api_router.post("/auth/register", response_model=User)
async def register_user(user: UserCreate):
    # Check if username already exists
    existing_user = await db.users.find_one({"username": user.username})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already registered"
        )
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": user.email})
    if existing_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Create new user
    hashed_password = get_password_hash(user.password)
    user_data = user.dict()
    del user_data["password"]
    
    # Set default permissions if not provided
    if not user_data.get("permissions"):
        user_data["permissions"] = get_default_permissions(user.role)
    
    user_in_db = UserInDB(**user_data, hashed_password=hashed_password)
    user_dict = user_in_db.dict()
    
    # Convert datetime objects to ISO strings for MongoDB
    if user_dict.get("created_at"):
        user_dict["created_at"] = user_dict["created_at"].isoformat()
    if user_dict.get("updated_at"):
        user_dict["updated_at"] = user_dict["updated_at"].isoformat()
    if user_dict.get("permissions") and hasattr(user_dict["permissions"], 'dict'):
        # Convert permissions to dict for MongoDB storage
        user_dict["permissions"] = user_dict["permissions"].dict()
    
    await db.users.insert_one(user_dict)
    
    # Return user without password
    return User(**user_data)

@api_router.post("/auth/login", response_model=Token)
async def login_user(user_credentials: UserLogin):
    user = await authenticate_user(user_credentials.username, user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Inactive user"
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    
    # Return user data without password
    user_data = User(**{k: v for k, v in user.dict().items() if k != "hashed_password"})
    
    return {"access_token": access_token, "token_type": "bearer", "user": user_data}

@api_router.get("/auth/me", response_model=User)
async def read_users_me(current_user: UserInDB = Depends(get_current_active_user)):
    # Return user data without password
    return User(**{k: v for k, v in current_user.dict().items() if k != "hashed_password"})

@api_router.post("/auth/logout")
async def logout_user():
    return {"message": "Successfully logged out"}

@api_router.get("/permissions/definitions")
async def get_permission_definitions(current_user: UserInDB = Depends(get_current_active_user)):
    """Get all available permissions with descriptions"""
    return {
        "permissions": {
            "medicines_view": "View medicines inventory",
            "medicines_add": "Add new medicines",
            "medicines_edit": "Edit medicine details",
            "medicines_delete": "Delete medicines",
            
            "patients_view": "View patient records",
            "patients_add": "Add new patients",
            "patients_edit": "Edit patient information",
            "patients_delete": "Delete patient records",
            
            "sales_view": "View sales records",
            "sales_add": "Create new sales",
            "sales_edit": "Edit sales records",
            "sales_delete": "Delete sales records",
            "sales_refund": "Process refunds",
            
            "doctors_view": "View doctor profiles",
            "doctors_add": "Add new doctors",
            "doctors_edit": "Edit doctor information",
            "doctors_delete": "Delete doctor profiles",
            
            "opd_view": "View OPD prescriptions",
            "opd_add": "Create new prescriptions",
            "opd_edit": "Edit prescriptions",
            "opd_delete": "Delete prescriptions",
            
            "analytics_view": "View analytics and reports",
            "analytics_export": "Export analytics data",
            
            "settings_view": "View system settings",
            "settings_edit": "Modify system settings",
            
            "users_view": "View user accounts",
            "users_add": "Create new users",
            "users_edit": "Edit user accounts",
            "users_delete": "Delete user accounts",
            
            "backup_create": "Create system backups",
            "backup_restore": "Restore from backups",
            "backup_delete": "Delete backup files"
        },
        "categories": {
            "medicines": ["medicines_view", "medicines_add", "medicines_edit", "medicines_delete"],
            "patients": ["patients_view", "patients_add", "patients_edit", "patients_delete"],
            "sales": ["sales_view", "sales_add", "sales_edit", "sales_delete", "sales_refund"],
            "doctors": ["doctors_view", "doctors_add", "doctors_edit", "doctors_delete"],
            "opd": ["opd_view", "opd_add", "opd_edit", "opd_delete"],
            "analytics": ["analytics_view", "analytics_export"],
            "settings": ["settings_view", "settings_edit"],
            "users": ["users_view", "users_add", "users_edit", "users_delete"],
            "backup": ["backup_create", "backup_restore", "backup_delete"]
        }
    }


# User Management APIs (Admin only)
@api_router.get("/users", response_model=List[User])
@require_permission("users_view")
async def get_users(current_user: UserInDB = Depends(get_current_active_user)):
    users = await db.users.find().to_list(1000)
    result = []
    for user in users:
        user_data = {k: v for k, v in user.items() if k != "hashed_password"}
        # Handle permissions field conversion
        if "permissions" in user_data and isinstance(user_data["permissions"], dict):
            user_data["permissions"] = Permission(**user_data["permissions"])
        elif "permissions" not in user_data:
            user_data["permissions"] = get_default_permissions(user.get("role", "staff"))
        result.append(User(**user_data))
    return result

@api_router.get("/users/{user_id}", response_model=User)
async def get_user_by_id(user_id: str, current_user: UserInDB = Depends(get_current_active_user)):
    if current_user.role not in ["admin", "manager"] and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return User(**{k: v for k, v in user.items() if k != "hashed_password"})

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_update: UserUpdate, current_user: UserInDB = Depends(get_current_active_user)):
    # Check permissions for editing users
    if not check_permission(current_user, "users_edit") and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to edit users"
        )
    
    # Only admin can change roles and permissions of other users
    if (user_update.role or user_update.permissions) and current_user.role != "admin" and current_user.id != user_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin can change user roles and permissions"
        )
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {k: v for k, v in user_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    # Convert permissions to dict for MongoDB storage
    if "permissions" in update_data and update_data["permissions"]:
        update_data["permissions"] = update_data["permissions"].dict()
    
    # If role is being changed and no permissions provided, set default permissions
    if "role" in update_data and "permissions" not in update_data:
        update_data["permissions"] = get_default_permissions(update_data["role"]).dict()
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    updated_user = await db.users.find_one({"id": user_id})
    
    # Convert permissions back to Permission object
    user_data = {k: v for k, v in updated_user.items() if k != "hashed_password"}
    if "permissions" in user_data and isinstance(user_data["permissions"], dict):
        user_data["permissions"] = Permission(**user_data["permissions"])
    
    return User(**user_data)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: UserInDB = Depends(get_current_active_user)):
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admin can delete users"
        )
    
    if current_user.id == user_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account"
        )
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Initialize default admin user (manual endpoint - auto-creation happens on startup)
@api_router.post("/auth/init-admin")
async def init_admin_user():
    """Manual admin user creation endpoint (admin is auto-created on startup)"""
    # Check if any admin user exists
    admin_exists = await db.users.find_one({"role": "admin"})
    if admin_exists:
        return {
            "message": "Admin user already exists",
            "username": "admin",
            "note": "Default password is 'admin123' if unchanged. Admin user is auto-created on server startup."
        }
    
    # Create default admin user (fallback if startup creation failed)
    admin_data = {
        "username": "admin",
        "email": "admin@medipos.com", 
        "full_name": "System Administrator",
        "role": "admin",
        "permissions": get_default_permissions("admin"),
        "is_active": True
    }
    
    hashed_password = get_password_hash("admin123")
    user_in_db = UserInDB(**admin_data, hashed_password=hashed_password)
    user_dict = user_in_db.dict()
    
    # Convert datetime objects to ISO strings for MongoDB
    if user_dict.get("created_at"):
        user_dict["created_at"] = user_dict["created_at"].isoformat()
    if user_dict.get("updated_at"):
        user_dict["updated_at"] = user_dict["updated_at"].isoformat()
    if user_dict.get("permissions") and hasattr(user_dict["permissions"], 'dict'):
        # Convert permissions to dict for MongoDB storage
        user_dict["permissions"] = user_dict["permissions"].dict()
    
    await db.users.insert_one(user_dict)
    
    return {
        "message": "Default admin user created successfully",
        "username": "admin",
        "password": "admin123",
        "note": "Please change the password after first login"
    }


# Medicine Management APIs
@api_router.post("/medicines", response_model=Medicine)
@require_permission("medicines_add")
async def create_medicine(medicine: MedicineCreate, current_user: UserInDB = Depends(get_current_active_user)):
    medicine_dict = medicine.dict()
    medicine_obj = Medicine(**medicine_dict)
    
    # Convert datetime objects to ISO strings for MongoDB
    medicine_dict = medicine_obj.dict()
    if medicine_dict.get("expiry_date"):
        medicine_dict["expiry_date"] = medicine_dict["expiry_date"].isoformat()
    if medicine_dict.get("created_at"):
        medicine_dict["created_at"] = medicine_dict["created_at"].isoformat()
    if medicine_dict.get("updated_at"):
        medicine_dict["updated_at"] = medicine_dict["updated_at"].isoformat()
        
    await db.medicines.insert_one(medicine_dict)
    return medicine_obj

@api_router.get("/medicines", response_model=List[Medicine])
@require_permission("medicines_view")
async def get_medicines(current_user: UserInDB = Depends(get_current_active_user)):
    medicines = await db.medicines.find().to_list(1000)
    return [Medicine(**medicine) for medicine in medicines]

@api_router.get("/medicines/low-stock")
async def get_low_stock_medicines():
    medicines = await db.medicines.find({
        "$expr": {"$lt": ["$stock_quantity", "$minimum_stock_level"]}
    }).to_list(1000)
    return [Medicine(**medicine) for medicine in medicines]

@api_router.get("/medicines/{medicine_id}", response_model=Medicine)
async def get_medicine(medicine_id: str):
    medicine = await db.medicines.find_one({"id": medicine_id})
    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")
    return Medicine(**medicine)

@api_router.put("/medicines/{medicine_id}", response_model=Medicine)
@require_permission("medicines_edit")
async def update_medicine(medicine_id: str, medicine_update: MedicineUpdate, current_user: UserInDB = Depends(get_current_active_user)):
    medicine = await db.medicines.find_one({"id": medicine_id})
    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    update_data = {k: v for k, v in medicine_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    await db.medicines.update_one({"id": medicine_id}, {"$set": update_data})
    updated_medicine = await db.medicines.find_one({"id": medicine_id})
    return Medicine(**updated_medicine)

@api_router.delete("/medicines/{medicine_id}")
@require_permission("medicines_delete")
async def delete_medicine(medicine_id: str, current_user: UserInDB = Depends(get_current_active_user)):
    result = await db.medicines.delete_one({"id": medicine_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Medicine not found")
    return {"message": "Medicine deleted successfully"}


# Patient Management APIs
@api_router.post("/patients", response_model=Patient)
async def create_patient(patient: PatientCreate):
    patient_dict = patient.dict()
    patient_obj = Patient(**patient_dict)
    await db.patients.insert_one(patient_obj.dict())
    return patient_obj

@api_router.get("/patients", response_model=List[Patient])
async def get_patients():
    patients = await db.patients.find().to_list(1000)
    return [Patient(**patient) for patient in patients]

@api_router.get("/patients/{patient_id}", response_model=Patient)
async def get_patient(patient_id: str):
    patient = await db.patients.find_one({"id": patient_id})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return Patient(**patient)

@api_router.put("/patients/{patient_id}", response_model=Patient)
async def update_patient(patient_id: str, patient_update: PatientUpdate):
    patient = await db.patients.find_one({"id": patient_id})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    update_data = {k: v for k, v in patient_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    await db.patients.update_one({"id": patient_id}, {"$set": update_data})
    updated_patient = await db.patients.find_one({"id": patient_id})
    return Patient(**updated_patient)

@api_router.delete("/patients/{patient_id}")
async def delete_patient(patient_id: str):
    result = await db.patients.delete_one({"id": patient_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Patient not found")
    return {"message": "Patient deleted successfully"}

@api_router.get("/patients/search/{query}")
async def search_patients(query: str):
    patients = await db.patients.find({
        "$or": [
            {"name": {"$regex": query, "$options": "i"}},
            {"phone": {"$regex": query, "$options": "i"}},
            {"email": {"$regex": query, "$options": "i"}}
        ]
    }).to_list(100)
    return [Patient(**patient) for patient in patients]


# Sales Management APIs
@api_router.post("/sales", response_model=Sale)
async def create_sale(sale: SaleCreate):
    # Validate and update stock quantities
    for item in sale.items:
        medicine = await db.medicines.find_one({"id": item.medicine_id})
        if not medicine:
            raise HTTPException(status_code=404, detail=f"Medicine {item.medicine_name} not found")
        
        if medicine["stock_quantity"] < item.quantity:
            raise HTTPException(
                status_code=400, 
                detail=f"Insufficient stock for {item.medicine_name}. Available: {medicine['stock_quantity']}"
            )
    
    # Create sale record
    sale_dict = sale.dict()
    sale_obj = Sale(**sale_dict)
    await db.sales.insert_one(sale_obj.dict())
    
    # Update stock quantities and create stock movements
    for item in sale.items:
        # Update medicine stock
        await db.medicines.update_one(
            {"id": item.medicine_id},
            {"$inc": {"stock_quantity": -item.quantity}}
        )
        
        # Create stock movement record
        stock_movement = StockMovement(
            medicine_id=item.medicine_id,
            medicine_name=item.medicine_name,
            transaction_type=TransactionType.SALE,
            quantity=-item.quantity,  # Negative for sale (stock reduction)
            unit_price=item.unit_price,
            total_value=-item.total_price,  # Negative for sale
            reference_id=sale_obj.id,
            notes=f"Sale to {sale.patient_name or 'Walk-in Customer'}"
        )
        await db.stock_movements.insert_one(stock_movement.dict())
    
    return sale_obj

@api_router.get("/sales", response_model=List[Sale])
async def get_sales():
    sales = await db.sales.find().sort("created_at", -1).to_list(1000)
    return [Sale(**sale) for sale in sales]

@api_router.get("/sales/today")
async def get_today_sales():
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.utcnow().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    sales = await db.sales.find({
        "created_at": {"$gte": today_start, "$lte": today_end}
    }).sort("created_at", -1).to_list(1000)
    
    total_sales = sum(sale["total_amount"] for sale in sales)
    
    return {
        "sales": [Sale(**sale) for sale in sales],
        "total_amount": total_sales,
        "total_transactions": len(sales)
    }

@api_router.get("/sales/{sale_id}", response_model=Sale)
async def get_sale(sale_id: str):
    sale = await db.sales.find_one({"id": sale_id})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    return Sale(**sale)


# Return/Refund Management APIs
@api_router.post("/returns", response_model=Return)
async def create_return(return_data: ReturnCreate):
    # Validate original sale exists
    original_sale = await db.sales.find_one({"id": return_data.original_sale_id})
    if not original_sale:
        raise HTTPException(status_code=404, detail="Original sale not found")
    
    # Validate return items against original sale
    for return_item in return_data.items:
        original_item = next((item for item in original_sale["items"] if item["medicine_id"] == return_item.medicine_id), None)
        if not original_item:
            raise HTTPException(status_code=400, detail=f"Medicine {return_item.medicine_name} was not in original sale")
        
        if return_item.quantity > original_item["quantity"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot return more {return_item.medicine_name} than originally purchased"
            )
    
    # Create return record
    return_obj = Return(**return_data.dict())
    await db.returns.insert_one(return_obj.dict())
    
    # Update stock quantities and create stock movements
    for item in return_data.items:
        # Update medicine stock (add back returned quantity)
        await db.medicines.update_one(
            {"id": item.medicine_id},
            {"$inc": {"stock_quantity": item.quantity}}
        )
        
        # Create stock movement record
        stock_movement = StockMovement(
            medicine_id=item.medicine_id,
            medicine_name=item.medicine_name,
            transaction_type=TransactionType.REFUND,
            quantity=item.quantity,  # Positive for return (stock addition)
            unit_price=item.unit_price,
            total_value=item.total_price,  # Positive for return
            reference_id=return_obj.id,
            notes=f"Return from sale #{return_data.original_sale_id[-8:]} - Reason: {return_data.reason or 'No reason provided'}"
        )
        await db.stock_movements.insert_one(stock_movement.dict())
    
    return return_obj

@api_router.get("/returns", response_model=List[Return])
async def get_returns():
    returns = await db.returns.find().sort("created_at", -1).to_list(1000)
    return [Return(**return_record) for return_record in returns]

@api_router.get("/returns/{return_id}", response_model=Return)
async def get_return(return_id: str):
    return_record = await db.returns.find_one({"id": return_id})
    if not return_record:
        raise HTTPException(status_code=404, detail="Return not found")
    return Return(**return_record)

@api_router.get("/sales/patient/{patient_id}")
async def get_patient_sales(patient_id: str):
    sales = await db.sales.find({"patient_id": patient_id}).sort("created_at", -1).to_list(100)
    return [Sale(**sale) for sale in sales]

@api_router.get("/returns/sale/{sale_id}")
async def get_sale_returns(sale_id: str):
    returns = await db.returns.find({"original_sale_id": sale_id}).sort("created_at", -1).to_list(1000)
    return [Return(**return_record) for return_record in returns]


# Stock Movement APIs
@api_router.get("/stock-movements", response_model=List[StockMovement])
async def get_stock_movements():
    movements = await db.stock_movements.find().sort("created_at", -1).to_list(1000)
    return [StockMovement(**movement) for movement in movements]

@api_router.get("/stock-movements/{medicine_id}")
async def get_medicine_stock_movements(medicine_id: str):
    movements = await db.stock_movements.find({"medicine_id": medicine_id}).sort("created_at", -1).to_list(1000)
    return [StockMovement(**movement) for movement in movements]


# Advanced Analytics APIs
@api_router.get("/analytics/comprehensive")
async def get_comprehensive_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    date_range: Optional[str] = None  # today, yesterday, this_week, this_month, custom
):
    """Get comprehensive analytics with date range filtering"""
    from datetime import datetime, timedelta
    
    # Parse date range
    now = datetime.utcnow()
    
    if date_range == "today":
        start_dt = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "yesterday":
        yesterday = now - timedelta(days=1)
        start_dt = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "this_week":
        start_dt = now - timedelta(days=now.weekday())
        start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "this_month":
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "custom" and start_date and end_date:
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        # Default to this month
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Get sales data for the period
    sales = await db.sales.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt}
    }).to_list(10000)
    
    # Get consultation data for the period
    consultations = await db.opd_prescriptions.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
    }).to_list(10000)
    
    # Calculate totals
    medicine_revenue = sum(sale["total_amount"] for sale in sales)
    consultation_revenue = sum(consultation.get("consultation_fee", 0) for consultation in consultations)
    total_revenue = medicine_revenue + consultation_revenue
    
    # Medicine analysis with profit margins
    medicine_analysis = {}
    total_cost = 0
    
    for sale in sales:
        for item in sale["items"]:
            med_id = item["medicine_id"]
            
            # Get medicine details for cost calculation
            medicine = await db.medicines.find_one({"id": med_id})
            if medicine:
                purchase_price = medicine.get("purchase_price", 0)
                cost = purchase_price * item["quantity"]
                profit = item["total_price"] - cost
                profit_margin = (profit / item["total_price"] * 100) if item["total_price"] > 0 else 0
                
                if med_id in medicine_analysis:
                    medicine_analysis[med_id]["quantity"] += item["quantity"]
                    medicine_analysis[med_id]["revenue"] += item["total_price"]
                    medicine_analysis[med_id]["cost"] += cost
                    medicine_analysis[med_id]["profit"] += profit
                else:
                    medicine_analysis[med_id] = {
                        "medicine_name": item["medicine_name"],
                        "manufacturer": medicine.get("manufacturer", "N/A"),
                        "generic_name": medicine.get("generic_name", "N/A"),
                        "quantity": item["quantity"],
                        "revenue": item["total_price"],
                        "cost": cost,
                        "profit": profit,
                        "purchase_price": purchase_price,
                        "selling_price": medicine.get("selling_price", 0)
                    }
                
                total_cost += cost
    
    # Calculate profit margins for each medicine
    for med_id in medicine_analysis:
        med = medicine_analysis[med_id]
        med["profit_margin"] = (med["profit"] / med["revenue"] * 100) if med["revenue"] > 0 else 0
    
    # Sort medicines by revenue
    top_medicines = sorted(medicine_analysis.values(), key=lambda x: x["revenue"], reverse=True)
    
    # Payment method breakdown
    payment_breakdown = {}
    for sale in sales:
        payment_method = sale.get("payment_method", "unknown")
        if payment_method in payment_breakdown:
            payment_breakdown[payment_method]["count"] += 1
            payment_breakdown[payment_method]["amount"] += sale["total_amount"]
        else:
            payment_breakdown[payment_method] = {
                "count": 1,
                "amount": sale["total_amount"]
            }
    
    # Manufacturer/Category breakdown
    manufacturer_breakdown = {}
    for med_data in medicine_analysis.values():
        manufacturer = med_data["manufacturer"]
        if manufacturer in manufacturer_breakdown:
            manufacturer_breakdown[manufacturer]["quantity"] += med_data["quantity"]
            manufacturer_breakdown[manufacturer]["revenue"] += med_data["revenue"]
            manufacturer_breakdown[manufacturer]["profit"] += med_data["profit"]
        else:
            manufacturer_breakdown[manufacturer] = {
                "quantity": med_data["quantity"],
                "revenue": med_data["revenue"],
                "profit": med_data["profit"]
            }
    
    return {
        "period": {
            "start_date": start_dt.isoformat(),
            "end_date": end_dt.isoformat(),
            "date_range": date_range or "custom"
        },
        "summary": {
            "total_revenue": total_revenue,
            "medicine_revenue": medicine_revenue,
            "consultation_revenue": consultation_revenue,
            "total_cost": total_cost,
            "total_profit": medicine_revenue - total_cost,
            "profit_margin": ((medicine_revenue - total_cost) / medicine_revenue * 100) if medicine_revenue > 0 else 0,
            "total_transactions": len(sales),
            "total_consultations": len(consultations)
        },
        "medicine_analysis": top_medicines,
        "payment_breakdown": payment_breakdown,
        "manufacturer_breakdown": manufacturer_breakdown
    }

@api_router.get("/analytics/monthly-comparison")
async def get_monthly_comparison(months: int = 12):
    """Get month-over-month sales comparison"""
    from datetime import datetime, timedelta
    import calendar
    
    now = datetime.utcnow()
    monthly_data = []
    
    for i in range(months):
        # Calculate month start and end
        month_date = now.replace(day=1) - timedelta(days=i*30)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Get last day of month
        last_day = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)
        
        # Get sales for this month
        sales = await db.sales.find({
            "created_at": {"$gte": month_start, "$lte": month_end}
        }).to_list(10000)
        
        # Get consultations for this month
        consultations = await db.opd_prescriptions.find({
            "created_at": {"$gte": month_start, "$lte": month_end},
            "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
        }).to_list(10000)
        
        medicine_revenue = sum(sale["total_amount"] for sale in sales)
        consultation_revenue = sum(consultation.get("consultation_fee", 0) for consultation in consultations)
        
        # Calculate costs and profits
        total_cost = 0
        for sale in sales:
            for item in sale["items"]:
                medicine = await db.medicines.find_one({"id": item["medicine_id"]})
                if medicine:
                    cost = medicine.get("purchase_price", 0) * item["quantity"]
                    total_cost += cost
        
        monthly_data.append({
            "month": month_start.strftime("%Y-%m"),
            "month_name": month_start.strftime("%B %Y"),
            "medicine_revenue": medicine_revenue,
            "consultation_revenue": consultation_revenue,
            "total_revenue": medicine_revenue + consultation_revenue,
            "total_cost": total_cost,
            "profit": medicine_revenue - total_cost,
            "transaction_count": len(sales),
            "consultation_count": len(consultations)
        })
    
    # Sort by month (newest first)
    monthly_data.sort(key=lambda x: x["month"], reverse=True)
    
    return {"monthly_data": monthly_data}

@api_router.get("/analytics/yearly-comparison")
async def get_yearly_comparison(years: int = 3):
    """Get year-over-year sales comparison"""
    from datetime import datetime
    
    now = datetime.utcnow()
    yearly_data = []
    
    for i in range(years):
        year = now.year - i
        year_start = datetime(year, 1, 1, 0, 0, 0, 0)
        year_end = datetime(year, 12, 31, 23, 59, 59, 999999)
        
        # Get sales for this year
        sales = await db.sales.find({
            "created_at": {"$gte": year_start, "$lte": year_end}
        }).to_list(10000)
        
        # Get consultations for this year
        consultations = await db.opd_prescriptions.find({
            "created_at": {"$gte": year_start, "$lte": year_end},
            "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
        }).to_list(10000)
        
        medicine_revenue = sum(sale["total_amount"] for sale in sales)
        consultation_revenue = sum(consultation.get("consultation_fee", 0) for consultation in consultations)
        
        # Calculate costs and profits
        total_cost = 0
        for sale in sales:
            for item in sale["items"]:
                medicine = await db.medicines.find_one({"id": item["medicine_id"]})
                if medicine:
                    cost = medicine.get("purchase_price", 0) * item["quantity"]
                    total_cost += cost
        
        yearly_data.append({
            "year": year,
            "medicine_revenue": medicine_revenue,
            "consultation_revenue": consultation_revenue,
            "total_revenue": medicine_revenue + consultation_revenue,
            "total_cost": total_cost,
            "profit": medicine_revenue - total_cost,
            "transaction_count": len(sales),
            "consultation_count": len(consultations)
        })
    
    return {"yearly_data": yearly_data}

@api_router.get("/analytics/kpis")
async def get_analytics_kpis(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    date_range: Optional[str] = None
):
    """Get Key Performance Indicators"""
    from datetime import datetime, timedelta
    
    # Parse date range (same logic as comprehensive analytics)
    now = datetime.utcnow()
    
    if date_range == "today":
        start_dt = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "yesterday":
        yesterday = now - timedelta(days=1)
        start_dt = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "this_week":
        start_dt = now - timedelta(days=now.weekday())
        start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "this_month":
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == "custom" and start_date and end_date:
        start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    else:
        start_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Current period data
    current_sales = await db.sales.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt}
    }).to_list(10000)
    
    current_consultations = await db.opd_prescriptions.find({
        "created_at": {"$gte": start_dt, "$lte": end_dt},
        "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
    }).to_list(10000)
    
    # Previous period for comparison
    period_length = end_dt - start_dt
    prev_start = start_dt - period_length
    prev_end = start_dt
    
    prev_sales = await db.sales.find({
        "created_at": {"$gte": prev_start, "$lte": prev_end}
    }).to_list(10000)
    
    prev_consultations = await db.opd_prescriptions.find({
        "created_at": {"$gte": prev_start, "$lte": prev_end},
        "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
    }).to_list(10000)
    
    # Calculate current metrics
    current_medicine_revenue = sum(sale["total_amount"] for sale in current_sales)
    current_consultation_revenue = sum(consultation.get("consultation_fee", 0) for consultation in current_consultations)
    current_total_revenue = current_medicine_revenue + current_consultation_revenue
    
    # Calculate previous metrics
    prev_medicine_revenue = sum(sale["total_amount"] for sale in prev_sales)
    prev_consultation_revenue = sum(consultation.get("consultation_fee", 0) for consultation in prev_consultations)  
    prev_total_revenue = prev_medicine_revenue + prev_consultation_revenue
    
    # Calculate growth percentages
    def calculate_growth(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return ((current - previous) / previous) * 100
    
    # Additional KPIs
    avg_transaction_value = current_medicine_revenue / len(current_sales) if current_sales else 0
    avg_consultation_fee = current_consultation_revenue / len(current_consultations) if current_consultations else 0
    
    # Customer metrics
    unique_patients = len(set(sale.get("patient_id") for sale in current_sales if sale.get("patient_id")))
    
    # Stock metrics
    total_medicines = await db.medicines.count_documents({})
    low_stock_count = await db.medicines.count_documents({
        "$expr": {"$lt": ["$stock_quantity", "$minimum_stock_level"]}
    })
    
    return {
        "period": {
            "start_date": start_dt.isoformat(),
            "end_date": end_dt.isoformat(),
            "date_range": date_range or "custom"
        },
        "revenue_kpis": {
            "total_revenue": {
                "current": current_total_revenue,
                "previous": prev_total_revenue,
                "growth": calculate_growth(current_total_revenue, prev_total_revenue)
            },
            "medicine_revenue": {
                "current": current_medicine_revenue,
                "previous": prev_medicine_revenue, 
                "growth": calculate_growth(current_medicine_revenue, prev_medicine_revenue)
            },
            "consultation_revenue": {
                "current": current_consultation_revenue,
                "previous": prev_consultation_revenue,
                "growth": calculate_growth(current_consultation_revenue, prev_consultation_revenue)
            }
        },
        "transaction_kpis": {
            "total_transactions": {
                "current": len(current_sales),
                "previous": len(prev_sales),
                "growth": calculate_growth(len(current_sales), len(prev_sales))
            },
            "total_consultations": {
                "current": len(current_consultations),
                "previous": len(prev_consultations),
                "growth": calculate_growth(len(current_consultations), len(prev_consultations))
            },
            "avg_transaction_value": avg_transaction_value,
            "avg_consultation_fee": avg_consultation_fee
        },
        "customer_kpis": {
            "unique_patients": unique_patients,
            "customer_retention": 0  # Can be enhanced later
        },
        "inventory_kpis": {
            "total_medicines": total_medicines,
            "low_stock_items": low_stock_count,
            "stock_turnover": 0  # Can be enhanced later
        }
    }

@api_router.get("/analytics/export-data")
async def get_export_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    date_range: Optional[str] = None,
    export_type: str = "comprehensive"  # comprehensive, medicine_analysis, sales_summary
):
    """Get data formatted for export (Excel/PDF)"""
    # Reuse the comprehensive analytics logic
    analytics_data = await get_comprehensive_analytics(start_date, end_date, date_range)
    
    if export_type == "medicine_analysis":
        return {
            "export_type": "medicine_analysis",
            "data": analytics_data["medicine_analysis"],
            "summary": analytics_data["summary"],
            "period": analytics_data["period"]
        }
    elif export_type == "sales_summary":
        return {
            "export_type": "sales_summary", 
            "data": {
                "summary": analytics_data["summary"],
                "payment_breakdown": analytics_data["payment_breakdown"],
                "manufacturer_breakdown": analytics_data["manufacturer_breakdown"]
            },
            "period": analytics_data["period"]
        }
    else:
        return {
            "export_type": "comprehensive",
            "data": analytics_data,
        }

# Analytics APIs
@api_router.get("/analytics/daily-sales-report")
async def get_daily_sales_report():
    """Get today's sales report with detailed payment method breakdown"""
    from datetime import datetime, timedelta
    
    # Get today's date range
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.utcnow().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Get today's sales
    today_sales = await db.sales.find({
        "created_at": {"$gte": today_start, "$lte": today_end}
    }).to_list(1000)
    
    # Initialize payment method breakdown
    payment_breakdown = {
        "cash": {"count": 0, "amount": 0.0, "transactions": []},
        "card": {"count": 0, "amount": 0.0, "transactions": []},
        "upi": {"count": 0, "amount": 0.0, "transactions": []},
        "credit": {"count": 0, "amount": 0.0, "transactions": []}
    }
    
    total_revenue = 0.0
    total_transactions = len(today_sales)
    
    # Process each sale
    for sale in today_sales:
        payment_method = sale.get("payment_method", "cash").lower()
        amount = sale.get("total_amount", 0.0)
        
        # Add to totals
        total_revenue += amount
        
        # Add to payment method breakdown
        if payment_method in payment_breakdown:
            payment_breakdown[payment_method]["count"] += 1
            payment_breakdown[payment_method]["amount"] += amount
            payment_breakdown[payment_method]["transactions"].append({
                "id": sale.get("id"),
                "patient_name": sale.get("patient_name", "Walk-in Customer"),
                "amount": amount,
                "items_count": len(sale.get("items", [])),
                "created_at": sale.get("created_at")
            })
    
    # Calculate percentages and averages
    for method in payment_breakdown:
        method_data = payment_breakdown[method]
        method_data["percentage"] = (method_data["amount"] / total_revenue * 100) if total_revenue > 0 else 0
        method_data["avg_transaction"] = (method_data["amount"] / method_data["count"]) if method_data["count"] > 0 else 0
    
    # Get hourly breakdown for today
    hourly_sales = {}
    for sale in today_sales:
        created_at = sale.get("created_at")
        if isinstance(created_at, str):
            try:
                created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            except:
                continue
        else:
            created_dt = created_at
        
        hour = created_dt.hour
        if hour not in hourly_sales:
            hourly_sales[hour] = {"count": 0, "amount": 0.0}
        
        hourly_sales[hour]["count"] += 1
        hourly_sales[hour]["amount"] += sale.get("total_amount", 0.0)
    
    # Sort hourly data
    sorted_hourly = [{"hour": hour, **data} for hour, data in sorted(hourly_sales.items())]
    
    return {
        "date": today_start.strftime("%Y-%m-%d"),
        "summary": {
            "total_revenue": total_revenue,
            "total_transactions": total_transactions,
            "avg_transaction_value": total_revenue / total_transactions if total_transactions > 0 else 0
        },
        "payment_breakdown": payment_breakdown,
        "hourly_breakdown": sorted_hourly,
        "period": {
            "start": today_start.isoformat(),
            "end": today_end.isoformat()
        }
    }

@api_router.get("/analytics/dashboard")
async def get_dashboard_analytics():
    # Get today's sales
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.utcnow().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    today_sales = await db.sales.find({
        "created_at": {"$gte": today_start, "$lte": today_end}
    }).to_list(1000)
    
    today_revenue = sum(sale["total_amount"] for sale in today_sales)
    today_transactions = len(today_sales)
    
    # Get total patients count
    total_patients = await db.patients.count_documents({})
    
    # Get total medicines count
    total_medicines = await db.medicines.count_documents({})
    
    # Get low stock count
    low_stock_count = await db.medicines.count_documents({
        "$expr": {"$lt": ["$stock_quantity", "$minimum_stock_level"]}
    })
    
    # Get top selling medicines (last 30 days)
    thirty_days_ago = datetime.utcnow().replace(day=1)  # Simplified to current month
    sales_last_30_days = await db.sales.find({
        "created_at": {"$gte": thirty_days_ago}
    }).to_list(1000)
    
    # Aggregate medicine sales
    medicine_sales = {}
    for sale in sales_last_30_days:
        for item in sale["items"]:
            med_id = item["medicine_id"]
            if med_id in medicine_sales:
                medicine_sales[med_id]["quantity"] += item["quantity"]
                medicine_sales[med_id]["revenue"] += item["total_price"]
            else:
                medicine_sales[med_id] = {
                    "medicine_name": item["medicine_name"],
                    "quantity": item["quantity"],
                    "revenue": item["total_price"]
                }
    
    # Sort by quantity sold
    top_medicines = sorted(medicine_sales.values(), key=lambda x: x["quantity"], reverse=True)[:5]
    
    return {
        "today_revenue": today_revenue,
        "today_transactions": today_transactions,
        "total_patients": total_patients,
        "total_medicines": total_medicines,
        "low_stock_count": low_stock_count,
        "top_selling_medicines": top_medicines
    }


# Doctor Management APIs
@api_router.post("/doctors", response_model=Doctor)
async def create_doctor(doctor: DoctorCreate):
    doctor_dict = doctor.dict()
    doctor_obj = Doctor(**doctor_dict)
    await db.doctors.insert_one(doctor_obj.dict())
    return doctor_obj

@api_router.get("/doctors", response_model=List[Doctor])
async def get_doctors():
    doctors = await db.doctors.find({"is_active": True}).to_list(1000)
    return [Doctor(**doctor) for doctor in doctors]

@api_router.get("/doctors/{doctor_id}", response_model=Doctor)
async def get_doctor(doctor_id: str):
    doctor = await db.doctors.find_one({"id": doctor_id})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    return Doctor(**doctor)

@api_router.put("/doctors/{doctor_id}", response_model=Doctor)
async def update_doctor(doctor_id: str, doctor_update: DoctorUpdate):
    doctor = await db.doctors.find_one({"id": doctor_id})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    update_data = {k: v for k, v in doctor_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow()
    
    await db.doctors.update_one({"id": doctor_id}, {"$set": update_data})
    updated_doctor = await db.doctors.find_one({"id": doctor_id})
    return Doctor(**updated_doctor)

@api_router.delete("/doctors/{doctor_id}")
async def delete_doctor(doctor_id: str):
    # Soft delete - set is_active to False
    result = await db.doctors.update_one({"id": doctor_id}, {"$set": {"is_active": False}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Doctor not found")
    return {"message": "Doctor deactivated successfully"}


# OPD Prescription APIs
@api_router.post("/opd-prescriptions", response_model=OPDPrescription)
async def create_opd_prescription(prescription: OPDPrescriptionCreate):
    # Validate doctor exists
    doctor = await db.doctors.find_one({"id": prescription.doctor_id, "is_active": True})
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    # Validate patient exists
    patient = await db.patients.find_one({"id": prescription.patient_id})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    prescription_dict = prescription.dict()
    prescription_obj = OPDPrescription(**prescription_dict)
    await db.opd_prescriptions.insert_one(prescription_obj.dict())
    return prescription_obj

@api_router.get("/opd-prescriptions", response_model=List[OPDPrescription])
async def get_opd_prescriptions():
    prescriptions = await db.opd_prescriptions.find().sort("created_at", -1).to_list(1000)
    return [OPDPrescription(**prescription) for prescription in prescriptions]

@api_router.get("/opd-prescriptions/{prescription_id}", response_model=OPDPrescription)
async def get_opd_prescription(prescription_id: str):
    prescription = await db.opd_prescriptions.find_one({"id": prescription_id})
    if not prescription:
        raise HTTPException(status_code=404, detail="Prescription not found")
    return OPDPrescription(**prescription)

@api_router.get("/opd-prescriptions/doctor/{doctor_id}")
async def get_doctor_prescriptions(doctor_id: str):
    prescriptions = await db.opd_prescriptions.find({"doctor_id": doctor_id}).sort("created_at", -1).to_list(1000)
    return [OPDPrescription(**prescription) for prescription in prescriptions]

@api_router.get("/opd-prescriptions/patient/{patient_id}")
async def get_patient_prescriptions(patient_id: str):
    prescriptions = await db.opd_prescriptions.find({"patient_id": patient_id}).sort("created_at", -1).to_list(1000)
    return [OPDPrescription(**prescription) for prescription in prescriptions]

# Generate OPD Prescription Print Format
@api_router.get("/opd-prescriptions/{prescription_id}/print")
async def get_prescription_print_data(prescription_id: str):
    prescription = await db.opd_prescriptions.find_one({"id": prescription_id})
    if not prescription:
        raise HTTPException(status_code=404, detail="Prescription not found")
    
    doctor = await db.doctors.find_one({"id": prescription["doctor_id"]})
    patient = await db.patients.find_one({"id": prescription["patient_id"]})
    
    if not doctor or not patient:
        raise HTTPException(status_code=404, detail="Doctor or Patient not found")
    
    return {
        "prescription": OPDPrescription(**prescription),
        "doctor": Doctor(**doctor),
        "patient": Patient(**patient),
        "print_date": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    }

# Sample data initialization
@api_router.post("/init-sample-data")
async def init_sample_data():
    # Check if data already exists
    medicine_count = await db.medicines.count_documents({})
    if medicine_count > 0:
        return {"message": "Sample data already exists"}
    
    # Sample medicines
    sample_medicines = [
        {
            "name": "Paracetamol 500mg",
            "generic_name": "Acetaminophen",
            "manufacturer": "ABC Pharma",
            "batch_number": "PCM001",
            "expiry_date": "2025-12-31",
            "purchase_price": 2.50,
            "selling_price": 5.00,
            "stock_quantity": 100,
            "minimum_stock_level": 20,
            "description": "Pain relief and fever reducer"
        },
        {
            "name": "Amoxicillin 250mg",
            "generic_name": "Amoxicillin",
            "manufacturer": "XYZ Pharmaceuticals",
            "batch_number": "AMX002",
            "expiry_date": "2025-10-31",
            "purchase_price": 8.00,
            "selling_price": 15.00,
            "stock_quantity": 50,
            "minimum_stock_level": 10,
            "description": "Antibiotic for bacterial infections"
        },
        {
            "name": "Cetirizine 10mg",
            "generic_name": "Cetirizine Hydrochloride",
            "manufacturer": "PQR Labs",
            "batch_number": "CTZ003",
            "expiry_date": "2026-03-31",
            "purchase_price": 3.00,
            "selling_price": 6.50,
            "stock_quantity": 75,
            "minimum_stock_level": 15,
            "description": "Antihistamine for allergies"
        },
        {
            "name": "Ibuprofen 400mg",
            "generic_name": "Ibuprofen",
            "manufacturer": "MNO Pharma",
            "batch_number": "IBU004",
            "expiry_date": "2025-08-31",
            "purchase_price": 4.50,
            "selling_price": 9.00,
            "stock_quantity": 30,
            "minimum_stock_level": 25,
            "description": "Anti-inflammatory pain reliever"
        },
        {
            "name": "Omeprazole 20mg",
            "generic_name": "Omeprazole",
            "manufacturer": "DEF Pharmaceuticals",
            "batch_number": "OME005",
            "expiry_date": "2025-11-30",
            "purchase_price": 6.00,
            "selling_price": 12.00,
            "stock_quantity": 40,
            "minimum_stock_level": 15,
            "description": "Proton pump inhibitor for acid reflux"
        }
    ]
    
    # Insert sample medicines
    medicines_to_insert = []
    for med_data in sample_medicines:
        med_data["expiry_date"] = datetime.strptime(med_data["expiry_date"], "%Y-%m-%d")
        medicine = Medicine(**med_data)
        # Convert date objects to strings for MongoDB
        medicine_dict = medicine.dict()
        if medicine_dict.get("expiry_date"):
            medicine_dict["expiry_date"] = medicine_dict["expiry_date"].isoformat()
        medicines_to_insert.append(medicine_dict)
    
    await db.medicines.insert_many(medicines_to_insert)
    
    # Sample patients
    sample_patients = [
        {
            "name": "John Smith",
            "phone": "+1234567890",
            "email": "john.smith@email.com",
            "address": "123 Main Street, City, State",
            "date_of_birth": "1985-05-15",
            "gender": "Male",
            "emergency_contact": "+1234567891",
            "medical_history": "Hypertension, Diabetes"
        },
        {
            "name": "Sarah Johnson",
            "phone": "+1234567892",
            "email": "sarah.j@email.com",
            "address": "456 Oak Avenue, City, State",
            "date_of_birth": "1990-08-22",
            "gender": "Female",
            "emergency_contact": "+1234567893",
            "medical_history": "Asthma"
        },
        {
            "name": "Michael Brown",
            "phone": "+1234567894",
            "email": "m.brown@email.com",
            "address": "789 Pine Road, City, State",
            "date_of_birth": "1978-12-10",
            "gender": "Male",
            "emergency_contact": "+1234567895",
            "medical_history": "None"
        }
    ]
    
    patients_to_insert = []
    for patient_data in sample_patients:
        patient_data["date_of_birth"] = datetime.strptime(patient_data["date_of_birth"], "%Y-%m-%d")
        patient = Patient(**patient_data)
        # Convert date objects to strings for MongoDB
        patient_dict = patient.dict()
        if patient_dict.get("date_of_birth"):
            patient_dict["date_of_birth"] = patient_dict["date_of_birth"].isoformat()
        patients_to_insert.append(patient_dict)
    
    await db.patients.insert_many(patients_to_insert)
    
    # Sample doctors
    sample_doctors = [
        {
            "name": "Dr. Aisha Khan",
            "specialization": "General Medicine",
            "qualification": "MD, MBBS",
            "license_number": "MED12345",
            "phone": "+1234567880",
            "email": "dr.aisha@clinic.com",
            "clinic_name": "City Health Clinic",
            "clinic_address": "123 Medical Avenue, City Center, State 12345",
            "consultation_fee": 150.00
        },
        {
            "name": "Dr. Rajesh Sharma",
            "specialization": "Pediatrics",
            "qualification": "MD Pediatrics, MBBS",
            "license_number": "PED67890",
            "phone": "+1234567881",
            "email": "dr.rajesh@childcare.com",
            "clinic_name": "Kids Care Pediatric Clinic",
            "clinic_address": "456 Children's Way, Medical District, State 12345",
            "consultation_fee": 180.00
        },
        {
            "name": "Dr. Maria Rodriguez",
            "specialization": "Dermatology",
            "qualification": "MD Dermatology, MBBS",
            "license_number": "DER13579",
            "phone": "+1234567882",
            "email": "dr.maria@skincare.com",
            "clinic_name": "Advanced Skin Solutions",
            "clinic_address": "789 Wellness Boulevard, Healthcare Plaza, State 12345",
            "consultation_fee": 200.00
        }
    ]
    
    doctors_to_insert = []
    for doctor_data in sample_doctors:
        doctor = Doctor(**doctor_data)
        doctors_to_insert.append(doctor.dict())
    
    await db.doctors.insert_many(doctors_to_insert)
    
    return {"message": "Sample data initialized successfully"}


# Settings Management APIs
@api_router.get("/settings")
async def get_settings():
    settings = await db.settings.find_one({})
    if not settings:
        # Return default settings if none exist
        default_settings = {
            "general": {
                "shop_name": "MediPOS Pharmacy",
                "shop_address": "123 Main Street, City, State, ZIP",
                "shop_phone": "+1-234-567-8900",
                "shop_email": "info@medipos.com",
                "shop_license": "PH-2024-001",
                "owner_name": "Pharmacy Owner",
                "gst_number": "",
                "currency": "USD",
                "currency_symbol": "$",
                "default_tax_rate": 10.0,
                "timezone": "UTC",
                "date_format": "YYYY-MM-DD",
                "time_format": "24",
                "decimal_places": 2,
                "language": "English",
                "backup_frequency": "daily",
                "auto_backup": True,
                "system_version": "1.0.0",
                "last_backup": None
            },
            "opd_paper": {
                "paper_size": "A4",
                "margin_top": 20,
                "margin_bottom": 20,
                "margin_left": 20,
                "margin_right": 20,
                "header_height": 80,
                "footer_height": 60,
                "line_height": 24,
                "font_size": 12,
                "font_family": "Arial",
                "show_logo": True,
                "logo_position": "left",
                "clinic_name_size": 18,
                "doctor_name_size": 14,
                "patient_info_size": 12,
                "prescription_area_lines": 15,
                "show_medical_history": True,
                "show_emergency_contact": False,
                "watermark_text": "",
                "print_instructions": "Please follow doctor's instructions carefully",
                "custom_html_enabled": False,
                "custom_html": "",
                "custom_css": ""
            },
            "printer": {
                "default_printer": "",
                "receipt_width": 80,
                "receipt_font_size": 10,
                "receipt_line_spacing": 1.2,
                "auto_print_receipts": False,
                "auto_print_prescriptions": False,
                "print_copies": 1,
                "paper_cutting": True,
                "cash_drawer": False,
                "barcode_format": "CODE128",
                "receipt_header": "MediPOS Pharmacy",
                "receipt_footer": "Thank you for your business!",
                "thermal_printer": False,
                "print_quality": "normal"
            },
            "telegram": {
                "enabled": False,
                "bot_token": "",
                "chat_id": "",
                "daily_report_time": "18:00",
                "include_revenue": True,
                "include_transactions": True,
                "include_top_medicines": True,
                "include_low_stock": True,
                "include_patient_count": True,
                "report_format": "detailed",
                "timezone": "UTC"
            },
            "alerts": {
                "low_stock_enabled": True,
                "low_stock_threshold": 10,
                "low_stock_check_frequency": "daily",
                "low_stock_notification_time": "09:00",
                "expiry_alert_enabled": True,
                "expiry_alert_days": 30,
                "expiry_check_frequency": "daily",
                "expiry_notification_time": "09:30",
                "telegram_alerts": False,
                "email_alerts": False,
                "system_notifications": True,
                "sound_notifications": False
            },
            "custom_templates": {}
        }
        return default_settings
    
    return {
        "general": settings.get("general", {}),
        "opd_paper": settings.get("opd_paper", {}),
        "printer": settings.get("printer", {}),
        "telegram": settings.get("telegram", {}),
        "alerts": settings.get("alerts", {}),
        "custom_templates": settings.get("custom_templates", {})
    }

@api_router.post("/settings")
async def save_settings(settings_data: SettingsUpdate):
    # Get existing settings or create new
    existing_settings = await db.settings.find_one({})
    
    if existing_settings:
        # Update existing settings
        update_data = {}
        if settings_data.general is not None:
            update_data["general"] = settings_data.general
        if settings_data.opd_paper is not None:
            update_data["opd_paper"] = settings_data.opd_paper
        if settings_data.printer is not None:
            update_data["printer"] = settings_data.printer
        if settings_data.telegram is not None:
            update_data["telegram"] = settings_data.telegram
        if settings_data.alerts is not None:
            update_data["alerts"] = settings_data.alerts
        if settings_data.custom_templates is not None:
            update_data["custom_templates"] = settings_data.custom_templates
        
        update_data["updated_at"] = datetime.utcnow()
        
        await db.settings.update_one({"_id": existing_settings["_id"]}, {"$set": update_data})
    else:
        # Create new settings
        new_settings = Settings(
            general=settings_data.general or {},
            opd_paper=settings_data.opd_paper or {},
            printer=settings_data.printer or {},
            telegram=settings_data.telegram or {},
            alerts=settings_data.alerts or {},
            custom_templates=settings_data.custom_templates or {}
        )
        await db.settings.insert_one(new_settings.dict())
    
    return {"message": "Settings saved successfully"}

@api_router.post("/test-telegram")
async def test_telegram_connection(telegram_data: dict):
    """Test Telegram bot connection"""
    import requests
    
    bot_token = telegram_data.get("bot_token")
    chat_id = telegram_data.get("chat_id")
    
    if not bot_token or not chat_id:
        return {"success": False, "error": "Bot token and chat ID are required"}
    
    try:
        # Test message
        test_message = """🧪 **MediPOS Test Message**
        
✅ Telegram connection is working properly!
📊 Daily sales reports will be sent to this chat.
⚙️ Configuration test completed successfully."""
        
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": test_message,
            "parse_mode": "Markdown"
        }
        
        response = requests.post(url, json=payload, timeout=10)
        
        if response.status_code == 200:
            return {"success": True, "message": "Test message sent successfully"}
        else:
            error_data = response.json()
            return {"success": False, "error": error_data.get("description", "Unknown error")}
            
    except Exception as e:
        return {"success": False, "error": str(e)}

@api_router.get("/system-status")
async def get_system_status():
    """Get system status and information"""
    import psutil
    import platform
    
    try:
        # Get system info
        system_info = {
            "system": platform.system(),
            "platform": platform.platform(),
            "architecture": platform.architecture()[0],
            "processor": platform.processor(),
            "python_version": platform.python_version(),
        }
        
        # Get memory info
        memory = psutil.virtual_memory()
        memory_info = {
            "total": memory.total,
            "available": memory.available,
            "percent": memory.percent,
            "used": memory.used,
            "free": memory.free
        }
        
        # Get disk info
        disk = psutil.disk_usage('/')
        disk_info = {
            "total": disk.total,
            "used": disk.used,
            "free": disk.free,
            "percent": (disk.used / disk.total) * 100
        }
        
        # Get CPU info
        cpu_info = {
            "count": psutil.cpu_count(),
            "percent": psutil.cpu_percent(interval=1),
            "frequency": psutil.cpu_freq()._asdict() if psutil.cpu_freq() else None
        }
        
        # Database status
        try:
            # Test database connection
            await db.command("ping")
            db_status = "Connected"
            
            # Get collection counts
            collections_info = {
                "medicines": await db.medicines.count_documents({}),
                "patients": await db.patients.count_documents({}),
                "doctors": await db.doctors.count_documents({}),
                "sales": await db.sales.count_documents({}),
                "returns": await db.returns.count_documents({}),
                "opd_prescriptions": await db.opd_prescriptions.count_documents({}),
                "stock_movements": await db.stock_movements.count_documents({})
            }
        except Exception as e:
            db_status = f"Error: {str(e)}"
            collections_info = {}
        
        return {
            "status": "operational",
            "timestamp": datetime.utcnow().isoformat(),
            "uptime": "N/A",  # Could be enhanced with actual uptime tracking
            "system": system_info,
            "memory": memory_info,
            "disk": disk_info,
            "cpu": cpu_info,
            "database": {
                "status": db_status,
                "collections": collections_info
            }
        }
        
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.utcnow().isoformat()
        }


async def generate_comprehensive_report():
    """Generate comprehensive daily report data"""
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.utcnow().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    # Get settings for currency symbol
    settings = await db.settings.find_one({}) or {}
    currency_symbol = settings.get("general", {}).get("currency_symbol", "$")
    expiry_alert_days = settings.get("alerts", {}).get("expiry_alert_days", 30)
    
    # Get today's sales data
    today_sales = await db.sales.find({
        "created_at": {"$gte": today_start, "$lte": today_end}
    }).to_list(1000)
    
    # Get today's OPD prescriptions with consultation fees
    today_prescriptions = await db.opd_prescriptions.find({
        "created_at": {"$gte": today_start, "$lte": today_end},
        "consultation_fee": {"$exists": True, "$ne": None, "$gt": 0}
    }).to_list(1000)
    
    # Calculate sales totals
    medicine_revenue = sum(sale["total_amount"] for sale in today_sales)
    consultation_revenue = sum(prescription.get("consultation_fee", 0) for prescription in today_prescriptions)
    total_revenue = medicine_revenue + consultation_revenue
    
    # Count transactions by payment method
    cash_transactions = len([s for s in today_sales if s.get("payment_method") == "cash"])
    upi_transactions = len([s for s in today_sales if s.get("payment_method") == "upi"])
    card_transactions = len([s for s in today_sales if s.get("payment_method") == "card"])
    credit_transactions = len([s for s in today_sales if s.get("payment_method") == "credit"])
    total_transactions = len(today_sales)
    
    # Calculate revenue by payment method
    cash_revenue = sum(sale["total_amount"] for sale in today_sales if sale.get("payment_method") == "cash")
    upi_revenue = sum(sale["total_amount"] for sale in today_sales if sale.get("payment_method") == "upi")
    card_revenue = sum(sale["total_amount"] for sale in today_sales if sale.get("payment_method") == "card")
    credit_revenue = sum(sale["total_amount"] for sale in today_sales if sale.get("payment_method") == "credit")
    
    # Get low stock medicines
    low_stock_medicines = await db.medicines.find({
        "$expr": {"$lt": ["$stock_quantity", "$minimum_stock_level"]}
    }).to_list(100)
    
    # Get medicines expiring soon
    expiry_alert_date = datetime.utcnow() + timedelta(days=expiry_alert_days)
    expiring_medicines = await db.medicines.find({
        "expiry_date": {
            "$exists": True,
            "$ne": None,
            "$lte": expiry_alert_date.isoformat()
        }
    }).to_list(100)
    
    # Filter medicines that are actually expired vs expiring soon
    today_date = datetime.utcnow()
    expired_medicines = []
    expiring_soon_medicines = []
    
    for medicine in expiring_medicines:
        if medicine.get("expiry_date"):
            try:
                expiry_date = datetime.fromisoformat(medicine["expiry_date"].replace("Z", "+00:00"))
                if expiry_date < today_date:
                    expired_medicines.append(medicine)
                else:
                    expiring_soon_medicines.append(medicine)
            except:
                # Handle different date formats
                continue
    
    return {
        "currency_symbol": currency_symbol,
        "total_revenue": total_revenue,
        "medicine_revenue": medicine_revenue,
        "consultation_revenue": consultation_revenue,
        "total_transactions": total_transactions,
        "consultation_count": len(today_prescriptions),
        "cash_transactions": cash_transactions,
        "cash_revenue": cash_revenue,
        "upi_transactions": upi_transactions,
        "upi_revenue": upi_revenue,
        "card_transactions": card_transactions,
        "card_revenue": card_revenue,
        "credit_transactions": credit_transactions,
        "credit_revenue": credit_revenue,
        "low_stock_medicines": low_stock_medicines,
        "expired_medicines": expired_medicines,
        "expiring_soon_medicines": expiring_soon_medicines,
        "expiry_alert_days": expiry_alert_days
    }

@api_router.post("/send-test-daily-report")
async def send_test_daily_report():
    """Send a comprehensive test daily report"""
    import requests
    
    # Get settings
    settings = await db.settings.find_one({})
    if not settings or not settings.get("telegram", {}).get("enabled"):
        raise HTTPException(status_code=400, detail="Telegram notifications not enabled")
    
    telegram_config = settings["telegram"]
    bot_token = telegram_config.get("bot_token")
    chat_id = telegram_config.get("chat_id")
    
    if not bot_token or not chat_id:
        raise HTTPException(status_code=400, detail="Telegram bot token or chat ID not configured")
    
    try:
        # Generate comprehensive report data
        report_data = await generate_comprehensive_report()
        
        # Create comprehensive report message
        report_message = f"""📊 **MediPOS Daily Report - TEST**
📅 Date: {datetime.utcnow().strftime('%Y-%m-%d')}

💰 **DAILY SALES SUMMARY**
├─ Total Revenue: {report_data['currency_symbol']}{report_data['total_revenue']:.2f}
├─ Medicine Sales: {report_data['currency_symbol']}{report_data['medicine_revenue']:.2f}
├─ Consultation Fees: {report_data['currency_symbol']}{report_data['consultation_revenue']:.2f}
└─ Total Transactions: {report_data['total_transactions']} sales + {report_data['consultation_count']} consultations

💳 **PAYMENT BREAKDOWN**
├─ 💵 Cash: {report_data['cash_transactions']} ({report_data['currency_symbol']}{report_data['cash_revenue']:.2f})
├─ 📱 UPI: {report_data['upi_transactions']} ({report_data['currency_symbol']}{report_data['upi_revenue']:.2f})
├─ 💳 Card: {report_data['card_transactions']} ({report_data['currency_symbol']}{report_data['card_revenue']:.2f})
└─ 🏷️ Credit: {report_data['credit_transactions']} ({report_data['currency_symbol']}{report_data['credit_revenue']:.2f})

⚠️ **LOW STOCK ALERT** ({len(report_data['low_stock_medicines'])} items)"""

        # Add low stock details
        if report_data['low_stock_medicines']:
            report_message += "\n"
            for medicine in report_data['low_stock_medicines'][:5]:  # Show first 5
                stock_qty = medicine.get('stock_quantity', 0)
                min_level = medicine.get('minimum_stock_level', 0)
                report_message += f"├─ {medicine['name']}: {stock_qty}/{min_level}\n"
            
            if len(report_data['low_stock_medicines']) > 5:
                report_message += f"└─ ...and {len(report_data['low_stock_medicines']) - 5} more items\n"
        else:
            report_message += "\n✅ All medicines are well stocked!\n"

        # Add expiry alerts
        expired_count = len(report_data['expired_medicines'])
        expiring_count = len(report_data['expiring_soon_medicines'])
        
        report_message += f"\n⏰ **EXPIRY ALERTS**"
        
        if expired_count > 0:
            report_message += f"\n🚨 **EXPIRED**: {expired_count} medicines"
            for medicine in report_data['expired_medicines'][:3]:  # Show first 3
                try:
                    expiry_date = datetime.fromisoformat(medicine["expiry_date"].replace("Z", "+00:00"))
                    report_message += f"\n├─ {medicine['name']}: {expiry_date.strftime('%Y-%m-%d')}"
                except:
                    report_message += f"\n├─ {medicine['name']}: Date parsing error"
        
        if expiring_count > 0:
            report_message += f"\n⚡ **EXPIRING SOON**: {expiring_count} medicines (within {report_data['expiry_alert_days']} days)"
            for medicine in report_data['expiring_soon_medicines'][:3]:  # Show first 3
                try:
                    expiry_date = datetime.fromisoformat(medicine["expiry_date"].replace("Z", "+00:00"))
                    days_until_expiry = (expiry_date - datetime.utcnow()).days
                    report_message += f"\n├─ {medicine['name']}: {days_until_expiry} days ({expiry_date.strftime('%Y-%m-%d')})"
                except:
                    report_message += f"\n├─ {medicine['name']}: Date parsing error"
        
        if expired_count == 0 and expiring_count == 0:
            report_message += "\n✅ No expiry concerns!"

        report_message += f"""

🧪 **This is a test report**
✅ Daily reports will be sent automatically at {telegram_config.get('daily_report_time', '18:00')}
📋 Configure alerts in Settings → Telegram"""
        
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": report_message,
            "parse_mode": "Markdown"
        }
        
        response = requests.post(url, json=payload, timeout=10)
        
        if response.status_code == 200:
            return {"success": True, "message": "Comprehensive test daily report sent successfully"}
        else:
            error_data = response.json()
            raise HTTPException(status_code=400, detail=error_data.get("description", "Failed to send message"))
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error sending test report: {str(e)}")


# Custom Template Management APIs
@api_router.post("/custom-templates", response_model=CustomTemplate)
async def create_custom_template(template: CustomTemplateCreate):
    """Create a new custom template"""
    template_dict = template.dict()
    template_obj = CustomTemplate(**template_dict)
    
    # Convert datetime objects to ISO strings for MongoDB
    template_data = template_obj.dict()
    if template_data.get("created_at"):
        template_data["created_at"] = template_data["created_at"].isoformat()
    if template_data.get("updated_at"):
        template_data["updated_at"] = template_data["updated_at"].isoformat()
    
    await db.custom_templates.insert_one(template_data)
    return template_obj

@api_router.get("/custom-templates", response_model=List[CustomTemplate])
async def get_custom_templates(category: Optional[str] = None, is_public: Optional[bool] = None):
    """Get all custom templates with optional filtering"""
    filter_query = {}
    
    if category:
        filter_query["category"] = category
    if is_public is not None:
        filter_query["is_public"] = is_public
    
    templates = await db.custom_templates.find(filter_query).sort("created_at", -1).to_list(1000)
    return [CustomTemplate(**template) for template in templates]

@api_router.get("/custom-templates/{template_id}", response_model=CustomTemplate)
async def get_custom_template(template_id: str):
    """Get a specific custom template by ID"""
    template = await db.custom_templates.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return CustomTemplate(**template)

@api_router.put("/custom-templates/{template_id}", response_model=CustomTemplate)
async def update_custom_template(template_id: str, template_update: CustomTemplateUpdate):
    """Update an existing custom template"""
    template = await db.custom_templates.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    update_data = {k: v for k, v in template_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    await db.custom_templates.update_one({"id": template_id}, {"$set": update_data})
    updated_template = await db.custom_templates.find_one({"id": template_id})
    return CustomTemplate(**updated_template)

@api_router.delete("/custom-templates/{template_id}")
async def delete_custom_template(template_id: str):
    """Delete a custom template"""
    result = await db.custom_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"message": "Template deleted successfully"}

@api_router.post("/custom-templates/{template_id}/duplicate", response_model=CustomTemplate)
async def duplicate_custom_template(template_id: str, new_name: Optional[str] = None):
    """Duplicate an existing custom template"""
    original_template = await db.custom_templates.find_one({"id": template_id})
    if not original_template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Create a new template based on the original
    new_template_data = {
        "name": new_name or f"{original_template['name']} (Copy)",
        "description": original_template.get("description", ""),
        "html": original_template["html"],
        "css": original_template["css"],
        "category": "custom",
        "is_public": False
    }
    
    new_template = CustomTemplate(**new_template_data)
    template_data = new_template.dict()
    if template_data.get("created_at"):
        template_data["created_at"] = template_data["created_at"].isoformat()
    if template_data.get("updated_at"):
        template_data["updated_at"] = template_data["updated_at"].isoformat()
    
    await db.custom_templates.insert_one(template_data)
    return new_template

@api_router.post("/custom-templates/import")
async def import_custom_template(template_data: dict):
    """Import a custom template from JSON data"""
    try:
        # Validate required fields
        if not all(key in template_data for key in ["name", "html", "css"]):
            raise HTTPException(status_code=400, detail="Missing required fields: name, html, css")
        
        # Create template from imported data
        template_create = CustomTemplateCreate(
            name=template_data["name"],
            description=template_data.get("description", ""),
            html=template_data["html"],
            css=template_data["css"],
            category=template_data.get("category", "custom"),
            is_public=template_data.get("is_public", False)
        )
        
        return await create_custom_template(template_create)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error importing template: {str(e)}")

@api_router.get("/custom-templates/{template_id}/export")
async def export_custom_template(template_id: str):
    """Export a custom template as JSON"""
    template = await db.custom_templates.find_one({"id": template_id})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Remove MongoDB-specific fields and internal IDs
    export_data = {
        "name": template["name"],
        "description": template.get("description", ""),
        "html": template["html"],
        "css": template["css"],
        "category": template.get("category", "custom"),
        "version": "1.0",
        "exported_at": datetime.utcnow().isoformat()
    }
    
    return export_data


# Backup & Restore Models and APIs
class BackupStatus(str, Enum):
    CREATING = "creating"
    COMPLETED = "completed"
    FAILED = "failed"
    RESTORING = "restoring"
    CORRUPTED = "corrupted"

class BackupType(str, Enum):
    MANUAL = "manual"
    SCHEDULED = "scheduled"
    AUTOMATIC = "automatic"

@api_router.get("/backup/directory-info")
async def get_backup_directory_info():
    """Get backup directory path information"""
    try:
        backup_dir = Path.cwd() / "backups"
        return {
            "backup_directory": str(backup_dir.absolute()),
            "exists": backup_dir.exists(),
            "platform": "Windows" if os.name == 'nt' else "Unix",
            "current_working_directory": str(Path.cwd()),
            "recommended_path": str(backup_dir.absolute())
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get directory info: {str(e)}")

class BackupInfo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    backup_type: BackupType = BackupType.MANUAL
    status: BackupStatus = BackupStatus.CREATING
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    checksum: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    database_collections: Optional[dict] = None
    app_version: Optional[str] = None
    created_by: Optional[str] = None
    error_message: Optional[str] = None

class BackupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    include_database: bool = True
    include_settings: bool = True
    include_app_files: bool = False
    create_archive: bool = True

class RestoreRequest(BaseModel):
    backup_id: str
    restore_database: bool = True
    restore_settings: bool = True
    force_restore: bool = False

# Backup Management APIs
@api_router.post("/backup/create")
async def create_backup(backup_request: BackupCreate):
    """Create a new backup"""
    try:
        # Create backup directory relative to current working directory (works on both Windows and Unix)
        backup_dir = Path.cwd() / "backups"
        backup_dir.mkdir(exist_ok=True)
        
        # Generate backup info
        backup_info = BackupInfo(
            name=backup_request.name,
            description=backup_request.description,
            backup_type=BackupType.MANUAL,
            status=BackupStatus.CREATING,
            app_version="1.0.0",
            created_by="system"
        )
        
        # Save backup info to database
        backup_dict = backup_info.dict()
        backup_dict["created_at"] = backup_dict["created_at"].isoformat()
        await db.backups.insert_one(backup_dict)
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        backup_folder = backup_dir / f"backup_{timestamp}_{backup_info.id[:8]}"
        backup_folder.mkdir(exist_ok=True)
        
        # Create backup data
        backup_data = {
            "backup_info": backup_info.dict(),
            "timestamp": timestamp,
            "collections": {}
        }
        
        try:
            # Backup database collections
            if backup_request.include_database:
                collections_info = {}
                
                # Backup medicines
                medicines = await db.medicines.find().to_list(10000)
                if medicines:
                    with open(backup_folder / "medicines.json", "w") as f:
                        json.dump(medicines, f, default=str, indent=2)
                    collections_info["medicines"] = len(medicines)
                
                # Backup patients
                patients = await db.patients.find().to_list(10000)
                if patients:
                    with open(backup_folder / "patients.json", "w") as f:
                        json.dump(patients, f, default=str, indent=2)
                    collections_info["patients"] = len(patients)
                
                # Backup doctors
                doctors = await db.doctors.find().to_list(10000)
                if doctors:
                    with open(backup_folder / "doctors.json", "w") as f:
                        json.dump(doctors, f, default=str, indent=2)
                    collections_info["doctors"] = len(doctors)
                
                # Backup sales
                sales = await db.sales.find().to_list(10000)
                if sales:
                    with open(backup_folder / "sales.json", "w") as f:
                        json.dump(sales, f, default=str, indent=2)
                    collections_info["sales"] = len(sales)
                
                # Backup OPD prescriptions
                prescriptions = await db.opd_prescriptions.find().to_list(10000)
                if prescriptions:
                    with open(backup_folder / "opd_prescriptions.json", "w") as f:
                        json.dump(prescriptions, f, default=str, indent=2)
                    collections_info["opd_prescriptions"] = len(prescriptions)
                
                # Backup stock movements
                stock_movements = await db.stock_movements.find().to_list(10000)
                if stock_movements:
                    with open(backup_folder / "stock_movements.json", "w") as f:
                        json.dump(stock_movements, f, default=str, indent=2)
                    collections_info["stock_movements"] = len(stock_movements)
                
                # Backup returns
                returns = await db.returns.find().to_list(10000)
                if returns:
                    with open(backup_folder / "returns.json", "w") as f:
                        json.dump(returns, f, default=str, indent=2)
                    collections_info["returns"] = len(returns)
                
                # Backup custom templates
                templates = await db.custom_templates.find().to_list(10000)
                if templates:
                    with open(backup_folder / "custom_templates.json", "w") as f:
                        json.dump(templates, f, default=str, indent=2)
                    collections_info["custom_templates"] = len(templates)
                
                backup_data["collections"] = collections_info
            
            # Backup settings
            if backup_request.include_settings:
                settings = await db.settings.find_one({})
                if settings:
                    # Remove MongoDB ObjectId
                    if "_id" in settings:
                        del settings["_id"]
                    with open(backup_folder / "settings.json", "w") as f:
                        json.dump(settings, f, default=str, indent=2)
            
            # Create backup metadata
            with open(backup_folder / "backup_metadata.json", "w") as f:
                json.dump(backup_data, f, default=str, indent=2)
            
            # Create archive if requested
            archive_path = None
            file_size = 0
            checksum = ""
            
            if backup_request.create_archive:
                archive_path = backup_dir / f"backup_{timestamp}_{backup_info.id[:8]}.tar.gz"
                
                # Create tar archive
                with tarfile.open(archive_path, "w:gz") as tar:
                    tar.add(backup_folder, arcname=f"backup_{timestamp}")
                
                file_size = archive_path.stat().st_size
                
                # Calculate checksum
                hash_md5 = hashlib.md5()
                with open(archive_path, "rb") as f:
                    for chunk in iter(lambda: f.read(4096), b""):
                        hash_md5.update(chunk)
                checksum = hash_md5.hexdigest()
                
                # Remove the folder since we have the archive
                shutil.rmtree(backup_folder)
            else:
                # Calculate folder size
                file_size = sum(f.stat().st_size for f in backup_folder.rglob('*') if f.is_file())
                archive_path = backup_folder
            
            # Update backup info
            completed_at = datetime.utcnow()
            await db.backups.update_one(
                {"id": backup_info.id},
                {"$set": {
                    "status": BackupStatus.COMPLETED.value,
                    "file_path": str(archive_path),
                    "file_size": file_size,
                    "checksum": checksum,
                    "completed_at": completed_at.isoformat(),
                    "database_collections": collections_info if backup_request.include_database else {}
                }}
            )
            
            return {
                "success": True,
                "backup_id": backup_info.id,
                "message": "Backup created successfully",
                "file_path": str(archive_path),
                "file_size": file_size,
                "collections": collections_info if backup_request.include_database else {}
            }
            
        except Exception as e:
            # Update backup status to failed
            await db.backups.update_one(
                {"id": backup_info.id},
                {"$set": {
                    "status": BackupStatus.FAILED.value,
                    "error_message": str(e)
                }}
            )
            raise HTTPException(status_code=500, detail=f"Backup creation failed: {str(e)}")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup initialization failed: {str(e)}")

@api_router.get("/backup/list")
async def list_backups():
    """List all available backups"""
    try:
        backups = await db.backups.find().sort("created_at", -1).to_list(100)
        return [BackupInfo(**backup) for backup in backups]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list backups: {str(e)}")

@api_router.get("/backup/{backup_id}")
async def get_backup_info(backup_id: str):
    """Get detailed information about a specific backup"""
    try:
        backup = await db.backups.find_one({"id": backup_id})
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        return BackupInfo(**backup)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get backup info: {str(e)}")

@api_router.post("/backup/restore")
async def restore_backup(restore_request: RestoreRequest):
    """Restore from a backup"""
    try:
        # Get backup info
        backup = await db.backups.find_one({"id": restore_request.backup_id})
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        if backup["status"] != BackupStatus.COMPLETED.value:
            raise HTTPException(status_code=400, detail="Backup is not completed or is corrupted")
        
        backup_info = BackupInfo(**backup)
        
        # Update backup status to restoring
        await db.backups.update_one(
            {"id": restore_request.backup_id},
            {"$set": {"status": BackupStatus.RESTORING.value}}
        )
        
        # Determine backup location
        backup_path = Path(backup_info.file_path)
        temp_dir = None
        
        if backup_path.suffix == '.gz':
            # Extract archive to temporary directory
            temp_dir = Path(tempfile.mkdtemp())
            with tarfile.open(backup_path, "r:gz") as tar:
                tar.extractall(temp_dir)
            # Find the backup folder inside
            backup_folder = next(temp_dir.rglob("backup_*"))
        else:
            backup_folder = backup_path
        
        try:
            restored_collections = {}
            
            # Restore database collections
            if restore_request.restore_database:
                # Restore medicines
                medicines_file = backup_folder / "medicines.json"
                if medicines_file.exists():
                    with open(medicines_file, "r") as f:
                        medicines_data = json.load(f)
                    if medicines_data:
                        if not restore_request.force_restore:
                            await db.medicines.delete_many({})
                        await db.medicines.insert_many(medicines_data)
                        restored_collections["medicines"] = len(medicines_data)
                
                # Restore patients
                patients_file = backup_folder / "patients.json"
                if patients_file.exists():
                    with open(patients_file, "r") as f:
                        patients_data = json.load(f)
                    if patients_data:
                        if not restore_request.force_restore:
                            await db.patients.delete_many({})
                        await db.patients.insert_many(patients_data)
                        restored_collections["patients"] = len(patients_data)
                
                # Restore doctors
                doctors_file = backup_folder / "doctors.json"
                if doctors_file.exists():
                    with open(doctors_file, "r") as f:
                        doctors_data = json.load(f)
                    if doctors_data:
                        if not restore_request.force_restore:
                            await db.doctors.delete_many({})
                        await db.doctors.insert_many(doctors_data)
                        restored_collections["doctors"] = len(doctors_data)
                
                # Restore sales
                sales_file = backup_folder / "sales.json"
                if sales_file.exists():
                    with open(sales_file, "r") as f:
                        sales_data = json.load(f)
                    if sales_data:
                        if not restore_request.force_restore:
                            await db.sales.delete_many({})
                        await db.sales.insert_many(sales_data)
                        restored_collections["sales"] = len(sales_data)
                
                # Restore OPD prescriptions
                prescriptions_file = backup_folder / "opd_prescriptions.json"
                if prescriptions_file.exists():
                    with open(prescriptions_file, "r") as f:
                        prescriptions_data = json.load(f)
                    if prescriptions_data:
                        if not restore_request.force_restore:
                            await db.opd_prescriptions.delete_many({})
                        await db.opd_prescriptions.insert_many(prescriptions_data)
                        restored_collections["opd_prescriptions"] = len(prescriptions_data)
                
                # Restore stock movements
                stock_movements_file = backup_folder / "stock_movements.json"
                if stock_movements_file.exists():
                    with open(stock_movements_file, "r") as f:
                        stock_movements_data = json.load(f)
                    if stock_movements_data:
                        if not restore_request.force_restore:
                            await db.stock_movements.delete_many({})
                        await db.stock_movements.insert_many(stock_movements_data)
                        restored_collections["stock_movements"] = len(stock_movements_data)
                
                # Restore returns
                returns_file = backup_folder / "returns.json"
                if returns_file.exists():
                    with open(returns_file, "r") as f:
                        returns_data = json.load(f)
                    if returns_data:
                        if not restore_request.force_restore:
                            await db.returns.delete_many({})
                        await db.returns.insert_many(returns_data)
                        restored_collections["returns"] = len(returns_data)
                
                # Restore custom templates
                templates_file = backup_folder / "custom_templates.json"
                if templates_file.exists():
                    with open(templates_file, "r") as f:
                        templates_data = json.load(f)
                    if templates_data:
                        if not restore_request.force_restore:
                            await db.custom_templates.delete_many({})
                        await db.custom_templates.insert_many(templates_data)
                        restored_collections["custom_templates"] = len(templates_data)
            
            # Restore settings
            if restore_request.restore_settings:
                settings_file = backup_folder / "settings.json"
                if settings_file.exists():
                    with open(settings_file, "r") as f:
                        settings_data = json.load(f)
                    if settings_data:
                        # Remove existing settings and insert restored one
                        await db.settings.delete_many({})
                        await db.settings.insert_one(settings_data)
                        restored_collections["settings"] = 1
            
            # Update backup status back to completed
            await db.backups.update_one(
                {"id": restore_request.backup_id},
                {"$set": {"status": BackupStatus.COMPLETED.value}}
            )
            
            return {
                "success": True,
                "message": "Restore completed successfully",
                "restored_collections": restored_collections,
                "backup_name": backup_info.name,
                "backup_date": backup_info.created_at
            }
            
        except Exception as e:
            # Update backup status back to completed (restore failed)
            await db.backups.update_one(
                {"id": restore_request.backup_id},
                {"$set": {"status": BackupStatus.COMPLETED.value}}
            )
            raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")
        
        finally:
            # Cleanup temporary directory
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)
                
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Restore operation failed: {str(e)}")

@api_router.delete("/backup/{backup_id}")
async def delete_backup(backup_id: str):
    """Delete a backup"""
    try:
        backup = await db.backups.find_one({"id": backup_id})
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        backup_info = BackupInfo(**backup)
        
        # Delete backup file/folder
        if backup_info.file_path:
            backup_path = Path(backup_info.file_path)
            if backup_path.exists():
                if backup_path.is_file():
                    backup_path.unlink()
                else:
                    shutil.rmtree(backup_path)
        
        # Delete backup record from database
        await db.backups.delete_one({"id": backup_id})
        
        return {
            "success": True,
            "message": "Backup deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete backup: {str(e)}")

@api_router.get("/backup/{backup_id}/verify")
async def verify_backup(backup_id: str):
    """Verify backup integrity"""
    try:
        backup = await db.backups.find_one({"id": backup_id})
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        backup_info = BackupInfo(**backup)
        
        if not backup_info.file_path:
            raise HTTPException(status_code=400, detail="Backup file path not found")
        
        backup_path = Path(backup_info.file_path)
        if not backup_path.exists():
            # Update status to corrupted
            await db.backups.update_one(
                {"id": backup_id},
                {"$set": {"status": BackupStatus.CORRUPTED.value}}
            )
            return {
                "success": False,
                "error": "Backup file not found",
                "status": "corrupted"
            }
        
        # Verify checksum if available
        if backup_info.checksum and backup_path.is_file():
            hash_md5 = hashlib.md5()
            with open(backup_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            calculated_checksum = hash_md5.hexdigest()
            
            if calculated_checksum != backup_info.checksum:
                await db.backups.update_one(
                    {"id": backup_id},
                    {"$set": {"status": BackupStatus.CORRUPTED.value}}
                )
                return {
                    "success": False,
                    "error": "Checksum mismatch - backup may be corrupted",
                    "status": "corrupted",
                    "expected_checksum": backup_info.checksum,
                    "calculated_checksum": calculated_checksum
                }
        
        # Basic file structure verification for archives
        if backup_path.suffix == '.gz':
            try:
                with tarfile.open(backup_path, "r:gz") as tar:
                    members = tar.getnames()
                    # Check if required files exist
                    has_metadata = any("backup_metadata.json" in member for member in members)
                    if not has_metadata:
                        return {
                            "success": False,
                            "error": "Backup metadata file not found",
                            "status": "corrupted"
                        }
            except Exception as e:
                await db.backups.update_one(
                    {"id": backup_id},
                    {"$set": {"status": BackupStatus.CORRUPTED.value}}
                )
                return {
                    "success": False,
                    "error": f"Cannot read backup archive: {str(e)}",
                    "status": "corrupted"
                }
        
        return {
            "success": True,
            "message": "Backup verification successful",
            "status": "verified",
            "file_size": backup_path.stat().st_size,
            "checksum_verified": backup_info.checksum is not None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Backup verification failed: {str(e)}")

@api_router.post("/backup/cleanup")
async def cleanup_old_backups(days_to_keep: int = 30):
    """Clean up old backups"""
    try:
        cutoff_date = datetime.utcnow() - timedelta(days=days_to_keep)
        
        # Find old backups
        old_backups = await db.backups.find({
            "created_at": {"$lt": cutoff_date.isoformat()}
        }).to_list(1000)
        
        deleted_count = 0
        for backup_data in old_backups:
            backup_info = BackupInfo(**backup_data)
            
            # Delete backup file/folder
            if backup_info.file_path:
                backup_path = Path(backup_info.file_path)
                if backup_path.exists():
                    if backup_path.is_file():
                        backup_path.unlink()
                    else:
                        shutil.rmtree(backup_path)
            
            # Delete backup record
            await db.backups.delete_one({"id": backup_info.id})
            deleted_count += 1
        
        return {
            "success": True,
            "message": f"Cleanup completed. Deleted {deleted_count} old backups",
            "deleted_count": deleted_count,
            "cutoff_date": cutoff_date.isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cleanup failed: {str(e)}")


# XLS Export/Import Models and APIs
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File
from typing import Dict, Any

class XLSExportRequest(BaseModel):
    collections: List[str] = ["medicines", "patients", "sales", "doctors", "opd_prescriptions", "returns", "settings"]
    date_range_start: Optional[datetime] = None
    date_range_end: Optional[datetime] = None
    include_system_data: bool = True

class XLSImportResult(BaseModel):
    success: bool
    message: str
    imported_counts: Dict[str, int] = {}
    errors: List[str] = []
    warnings: List[str] = []

def convert_datetime_for_excel(obj):
    """Convert datetime objects to Excel-compatible format"""
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {k: convert_datetime_for_excel(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_datetime_for_excel(item) for item in obj]
    return obj

@api_router.post("/export/xls")
async def export_data_to_xls(export_request: XLSExportRequest):
    """Export selected data collections to XLS format"""
    try:
        workbook = Workbook()
        # Keep the default sheet initially, we'll remove it at the end if we have other sheets
        default_sheet = workbook.active
        default_sheet.title = "Export_Info"
        
        exported_collections = {}
        sheets_created = 0
        
        # Export Medicines
        if "medicines" in export_request.collections:
            medicines = await db.medicines.find().to_list(10000)
            if medicines:
                sheet = workbook.create_sheet("Medicines")
                
                # Headers
                headers = ["ID", "Name", "Generic Name", "Manufacturer", "Batch Number", 
                          "Expiry Date", "Purchase Price", "Selling Price", "Stock Quantity", 
                          "Minimum Stock Level", "Description", "Created At", "Updated At"]
                sheet.append(headers)
                
                # Data
                for medicine in medicines:
                    row = [
                        medicine.get("id", ""),
                        medicine.get("name", ""),
                        medicine.get("generic_name", ""),
                        medicine.get("manufacturer", ""),
                        medicine.get("batch_number", ""),
                        medicine.get("expiry_date", ""),
                        medicine.get("purchase_price", 0),
                        medicine.get("selling_price", 0),
                        medicine.get("stock_quantity", 0),
                        medicine.get("minimum_stock_level", 0),
                        medicine.get("description", ""),
                        medicine.get("created_at", ""),
                        medicine.get("updated_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["medicines"] = len(medicines)
                sheets_created += 1
        
        # Export Patients
        if "patients" in export_request.collections:
            patients = await db.patients.find().to_list(10000)
            if patients:
                sheet = workbook.create_sheet("Patients")
                
                headers = ["ID", "Name", "Phone", "Email", "Address", "Date of Birth", 
                          "Gender", "Emergency Contact", "Medical History", "Created At", "Updated At"]
                sheet.append(headers)
                
                for patient in patients:
                    row = [
                        patient.get("id", ""),
                        patient.get("name", ""),
                        patient.get("phone", ""),
                        patient.get("email", ""),
                        patient.get("address", ""),
                        patient.get("date_of_birth", ""),
                        patient.get("gender", ""),
                        patient.get("emergency_contact", ""),
                        patient.get("medical_history", ""),
                        patient.get("created_at", ""),
                        patient.get("updated_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["patients"] = len(patients)
                sheets_created += 1
        
        # Export Sales
        if "sales" in export_request.collections:
            query = {}
            if export_request.date_range_start or export_request.date_range_end:
                date_filter = {}
                if export_request.date_range_start:
                    date_filter["$gte"] = export_request.date_range_start
                if export_request.date_range_end:
                    date_filter["$lte"] = export_request.date_range_end
                query["created_at"] = date_filter
            
            sales = await db.sales.find(query).to_list(10000)
            if sales:
                sheet = workbook.create_sheet("Sales")
                
                headers = ["ID", "Patient ID", "Patient Name", "Items", "Subtotal", 
                          "Tax Amount", "Discount Amount", "Total Amount", "Payment Method", "Created At"]
                sheet.append(headers)
                
                for sale in sales:
                    items_str = "; ".join([f"{item.get('medicine_name', '')} ({item.get('quantity', 0)}x{item.get('unit_price', 0)})" 
                                         for item in sale.get("items", [])])
                    row = [
                        sale.get("id", ""),
                        sale.get("patient_id", ""),
                        sale.get("patient_name", ""),
                        items_str,
                        sale.get("subtotal", 0),
                        sale.get("tax_amount", 0),
                        sale.get("discount_amount", 0),
                        sale.get("total_amount", 0),
                        sale.get("payment_method", ""),
                        sale.get("created_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["sales"] = len(sales)
                sheets_created += 1
        
        # Export Doctors
        if "doctors" in export_request.collections:
            doctors = await db.doctors.find().to_list(1000)
            if doctors:
                sheet = workbook.create_sheet("Doctors")
                
                headers = ["ID", "Name", "Specialization", "Qualification", "License Number", 
                          "Phone", "Email", "Clinic Name", "Clinic Address", "Consultation Fee", 
                          "Is Active", "Created At", "Updated At"]
                sheet.append(headers)
                
                for doctor in doctors:
                    row = [
                        doctor.get("id", ""),
                        doctor.get("name", ""),
                        doctor.get("specialization", ""),
                        doctor.get("qualification", ""),
                        doctor.get("license_number", ""),
                        doctor.get("phone", ""),
                        doctor.get("email", ""),
                        doctor.get("clinic_name", ""),
                        doctor.get("clinic_address", ""),
                        doctor.get("consultation_fee", 0),
                        doctor.get("is_active", True),
                        doctor.get("created_at", ""),
                        doctor.get("updated_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["doctors"] = len(doctors)
                sheets_created += 1
        
        # Export OPD Prescriptions
        if "opd_prescriptions" in export_request.collections:
            query = {}
            if export_request.date_range_start or export_request.date_range_end:
                date_filter = {}
                if export_request.date_range_start:
                    date_filter["$gte"] = export_request.date_range_start
                if export_request.date_range_end:
                    date_filter["$lte"] = export_request.date_range_end
                query["created_at"] = date_filter
            
            prescriptions = await db.opd_prescriptions.find(query).to_list(10000)
            if prescriptions:
                sheet = workbook.create_sheet("OPD_Prescriptions")
                
                headers = ["ID", "Doctor ID", "Patient ID", "Date", "Consultation Fee", 
                          "Prescription Notes", "Next Visit Date", "Created At"]
                sheet.append(headers)
                
                for prescription in prescriptions:
                    row = [
                        prescription.get("id", ""),
                        prescription.get("doctor_id", ""),
                        prescription.get("patient_id", ""),
                        prescription.get("date", ""),
                        prescription.get("consultation_fee", 0),
                        prescription.get("prescription_notes", ""),
                        prescription.get("next_visit_date", ""),
                        prescription.get("created_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["opd_prescriptions"] = len(prescriptions)
                sheets_created += 1
        
        # Export Returns
        if "returns" in export_request.collections:
            returns = await db.returns.find().to_list(10000)
            if returns:
                sheet = workbook.create_sheet("Returns")
                
                headers = ["ID", "Original Sale ID", "Patient ID", "Patient Name", "Items", 
                          "Subtotal", "Tax Amount", "Discount Amount", "Total Amount", 
                          "Reason", "Refund Method", "Created At"]
                sheet.append(headers)
                
                for return_item in returns:
                    items_str = "; ".join([f"{item.get('medicine_name', '')} ({item.get('quantity', 0)}x{item.get('unit_price', 0)})" 
                                         for item in return_item.get("items", [])])
                    row = [
                        return_item.get("id", ""),
                        return_item.get("original_sale_id", ""),
                        return_item.get("patient_id", ""),
                        return_item.get("patient_name", ""),
                        items_str,
                        return_item.get("subtotal", 0),
                        return_item.get("tax_amount", 0),
                        return_item.get("discount_amount", 0),
                        return_item.get("total_amount", 0),
                        return_item.get("reason", ""),
                        return_item.get("refund_method", ""),
                        return_item.get("created_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["returns"] = len(returns)
                sheets_created += 1
        
        # Export Settings
        if "settings" in export_request.collections and export_request.include_system_data:
            settings = await db.settings.find().to_list(100)
            if settings:
                sheet = workbook.create_sheet("Settings")
                
                headers = ["ID", "General", "OPD Paper", "Printer", "Telegram", "Alerts", 
                          "Custom Templates", "Created At", "Updated At"]
                sheet.append(headers)
                
                for setting in settings:
                    row = [
                        setting.get("id", ""),
                        str(setting.get("general", {})),
                        str(setting.get("opd_paper", {})),
                        str(setting.get("printer", {})),
                        str(setting.get("telegram", {})),
                        str(setting.get("alerts", {})),
                        str(setting.get("custom_templates", {})),
                        setting.get("created_at", ""),
                        setting.get("updated_at", "")
                    ]
                    sheet.append(row)
                
                exported_collections["settings"] = len(settings)
                sheets_created += 1
        
        # Add export information to the default sheet
        if sheets_created > 0:
            # Add export summary to the info sheet
            default_sheet.append(["MediPOS Data Export Summary"])
            default_sheet.append(["Export Date:", datetime.utcnow().isoformat()])
            default_sheet.append([""])
            default_sheet.append(["Exported Collections:"])
            for collection, count in exported_collections.items():
                default_sheet.append([collection.replace('_', ' ').title(), count])
        else:
            # If no data sheets were created, add a message
            default_sheet.append(["No data found for the selected collections"])
            default_sheet.append(["Export Date:", datetime.utcnow().isoformat()])
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"medipos_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@api_router.post("/import/xls", response_model=XLSImportResult)
async def import_data_from_xls(file: UploadFile = File(...)):
    """Import data from XLS file with merge capability"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Read file content
        content = await file.read()
        try:
            workbook = load_workbook(BytesIO(content))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file format: {str(e)}")
        
        imported_counts = {}
        errors = []
        warnings = []
        
        # Import Medicines
        if "Medicines" in workbook.sheetnames:
            sheet = workbook["Medicines"]
            medicines_imported = 0
            
            # Skip header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:  # Skip if no ID or Name
                        continue
                    
                    # Check if medicine exists
                    existing = await db.medicines.find_one({"id": row[0]})
                    
                    medicine_data = {
                        "id": row[0],
                        "name": row[1],
                        "generic_name": row[2] or None,
                        "manufacturer": row[3] or None,
                        "batch_number": row[4] or None,
                        "expiry_date": row[5] or None,
                        "purchase_price": float(row[6]) if row[6] else 0,
                        "selling_price": float(row[7]) if row[7] else 0,
                        "stock_quantity": int(row[8]) if row[8] else 0,
                        "minimum_stock_level": int(row[9]) if row[9] else 10,
                        "description": row[10] or None,
                        "updated_at": datetime.utcnow()
                    }
                    
                    if existing:
                        # Update existing (merge)
                        await db.medicines.update_one(
                            {"id": row[0]}, 
                            {"$set": medicine_data}
                        )
                        warnings.append(f"Updated existing medicine: {row[1]}")
                    else:
                        # Create new
                        medicine_data["created_at"] = datetime.utcnow()
                        await db.medicines.insert_one(medicine_data)
                    
                    medicines_imported += 1
                    
                except Exception as e:
                    errors.append(f"Medicine row error: {str(e)}")
            
            imported_counts["medicines"] = medicines_imported
        
        # Import Patients
        if "Patients" in workbook.sheetnames:
            sheet = workbook["Patients"]
            patients_imported = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:  # Skip if no ID or Name
                        continue
                    
                    existing = await db.patients.find_one({"id": row[0]})
                    
                    patient_data = {
                        "id": row[0],
                        "name": row[1],
                        "phone": row[2] or None,
                        "email": row[3] or None,
                        "address": row[4] or None,
                        "date_of_birth": row[5] or None,
                        "gender": row[6] or None,
                        "emergency_contact": row[7] or None,
                        "medical_history": row[8] or None,
                        "updated_at": datetime.utcnow()
                    }
                    
                    if existing:
                        await db.patients.update_one(
                            {"id": row[0]}, 
                            {"$set": patient_data}
                        )
                        warnings.append(f"Updated existing patient: {row[1]}")
                    else:
                        patient_data["created_at"] = datetime.utcnow()
                        await db.patients.insert_one(patient_data)
                    
                    patients_imported += 1
                    
                except Exception as e:
                    errors.append(f"Patient row error: {str(e)}")
            
            imported_counts["patients"] = patients_imported
        
        # Import Doctors
        if "Doctors" in workbook.sheetnames:
            sheet = workbook["Doctors"]
            doctors_imported = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    if not row[0] or not row[1]:  # Skip if no ID or Name
                        continue
                    
                    existing = await db.doctors.find_one({"id": row[0]})
                    
                    doctor_data = {
                        "id": row[0],
                        "name": row[1],
                        "specialization": row[2] or "",
                        "qualification": row[3] or "",
                        "license_number": row[4] or "",
                        "phone": row[5] or None,
                        "email": row[6] or None,
                        "clinic_name": row[7] or None,
                        "clinic_address": row[8] or None,
                        "consultation_fee": float(row[9]) if row[9] else None,
                        "is_active": bool(row[10]) if row[10] is not None else True,
                        "updated_at": datetime.utcnow()
                    }
                    
                    if existing:
                        await db.doctors.update_one(
                            {"id": row[0]}, 
                            {"$set": doctor_data}
                        )
                        warnings.append(f"Updated existing doctor: {row[1]}")
                    else:
                        doctor_data["created_at"] = datetime.utcnow()
                        await db.doctors.insert_one(doctor_data)
                    
                    doctors_imported += 1
                    
                except Exception as e:
                    errors.append(f"Doctor row error: {str(e)}")
            
            imported_counts["doctors"] = doctors_imported
        
        # Note: Sales, Returns, and OPD Prescriptions are typically not imported
        # as they represent historical transactions, but we can add them if needed
        
        result = XLSImportResult(
            success=len(errors) == 0,
            message=f"Import completed. Imported: {sum(imported_counts.values())} records",
            imported_counts=imported_counts,
            errors=errors,
            warnings=warnings
        )
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# Startup event to create default admin user
async def create_default_admin():
    """Create default admin user if no admin exists"""
    try:
        # Check if any admin user exists
        admin_exists = await db.users.find_one({"role": "admin"})
        if admin_exists:
            logger.info("Admin user already exists. Skipping default admin creation.")
            return
        
        # Create default admin user
        admin_data = {
            "username": "admin",
            "email": "admin@medipos.com", 
            "full_name": "System Administrator",
            "role": "admin",
            "permissions": get_default_permissions("admin"),
            "is_active": True
        }
        
        hashed_password = get_password_hash("admin123")
        user_in_db = UserInDB(**admin_data, hashed_password=hashed_password)
        user_dict = user_in_db.dict()
        
        # Convert datetime objects to ISO strings for MongoDB
        if user_dict.get("created_at"):
            user_dict["created_at"] = user_dict["created_at"].isoformat()
        if user_dict.get("updated_at"):
            user_dict["updated_at"] = user_dict["updated_at"].isoformat()
        if user_dict.get("permissions") and hasattr(user_dict["permissions"], 'dict'):
            # Convert permissions to dict for MongoDB storage
            user_dict["permissions"] = user_dict["permissions"].dict()
        
        await db.users.insert_one(user_dict)
        
        logger.info("✅ Default admin user created successfully!")
        logger.info("   Username: admin")
        logger.info("   Password: admin123")
        logger.info("   ⚠️  Please change the default password after first login!")
        
    except Exception as e:
        logger.error(f"Failed to create default admin user: {str(e)}")

@app.on_event("startup")
async def startup_event():
    """Application startup event"""
    logger.info("🚀 Starting MediPOS Backend Server...")
    logger.info("📊 Initializing database connection...")
    
    # Test database connection
    try:
        # Ping the database
        await client.admin.command('ping')
        logger.info("✅ Database connection established successfully!")
        
        # Create default admin user
        await create_default_admin()
        
        # Log startup completion
        logger.info("🎉 MediPOS Backend Server started successfully!")
        logger.info("📚 API Documentation available at: /docs")
        logger.info("🔐 Default Admin Login: admin/admin123")
        
    except Exception as e:
        logger.error(f"❌ Database connection failed: {str(e)}")
        logger.error("⚠️  Please ensure MongoDB is running and accessible")

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "server:app",
        host=os.environ.get("HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", 8001)),
        reload=True
    )